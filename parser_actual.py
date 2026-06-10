# ACTUAL VERSION: v17 (source: trendagent_parser_17.py)
#!/usr/bin/env python3
"""
Парсер TrendAgent — квартиры + планировки + рендеры ЖК
"""

import os
import time
import shutil
import requests
import re
import json
import boto3
from botocore.exceptions import BotoCoreError, ClientError
from PIL import Image
import io
from datetime import datetime
from pathlib import Path
from dotenv import load_dotenv

load_dotenv()  # читаем .env из папки скрипта

from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from bs4 import BeautifulSoup
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Журнал состояния проекта (project_log.py должен лежать рядом)
try:
    from project_log import write_run, update_state
except ImportError:
    # Если модуль ещё не подложен — заглушки чтобы скрипт работал
    def write_run(*args, **kwargs): pass
    def update_state(*args, **kwargs): pass

# ──────────────────────────────────────────────
# НАСТРОЙКИ
# ──────────────────────────────────────────────
BASE_URL  = "https://spb.trendagent.ru/objects/table/?apartments-balcony%5B0%5D=612de4576984843478489004&apartments-level_type%5B0%5D=65b8bfec34b5a527f81042ad&apartments-level_type%5B1%5D=65b8bfec34b5a527f81042ae&apartments-level_type%5B2%5D=65b8bfec34b5a527f81042af&apartments-iscityregion=true&apartments-floor_first=false&apartments-reservation=true&"
LIST_URL  = "https://spb.trendagent.ru/objects/list/?apartments-balcony%5B0%5D=612de4576984843478489004&apartments-level_type%5B0%5D=65b8bfec34b5a527f81042ad&apartments-level_type%5B1%5D=65b8bfec34b5a527f81042ae&apartments-level_type%5B2%5D=65b8bfec34b5a527f81042af&apartments-floor_first=false&apartments-iscityregion=true&apartments-reservation=true"

SCRIPT_DIR     = Path(__file__).parent
CHROME_PROFILE = SCRIPT_DIR / "chrome_profile"
LAYOUTS_DIR    = SCRIPT_DIR / "layouts"
RENDERS_DIR    = SCRIPT_DIR / "renders"

# ──────────────────────────────────────────────
# ЯНДЕКС OBJECT STORAGE
# ──────────────────────────────────────────────
JK_DATA_FILE   = SCRIPT_DIR / "jk_data.json"
S3_JK_DATA_KEY = "terrace/jk_data.json"

S3_ACCESS_KEY  = os.getenv("S3_ACCESS_KEY")
S3_SECRET_KEY  = os.getenv("S3_SECRET_KEY")
S3_BUCKET      = os.getenv("S3_BUCKET", "royaltyplace")
S3_PREFIX      = os.getenv("S3_PREFIX", "terrace/renders")

# ── Флаг: перезалить все планировки в WebP (не скачивать заново, только конвертировать из layouts/) ──
# Поставь True если планировки есть локально но не отображаются на сайте
FORCE_REUPLOAD_LAYOUTS = False
S3_ENDPOINT    = os.getenv("S3_ENDPOINT", "https://storage.yandexcloud.net")

if not S3_ACCESS_KEY or not S3_SECRET_KEY:
    raise RuntimeError("❌ S3_ACCESS_KEY / S3_SECRET_KEY не найдены — проверь файл .env")


def get_s3_client():
    return boto3.client(
        "s3",
        endpoint_url=S3_ENDPOINT,
        aws_access_key_id=S3_ACCESS_KEY,
        aws_secret_access_key=S3_SECRET_KEY,
        region_name="ru-central1",
    )


def compress_image(filepath: Path, max_width: int = 1200, quality: int = 85) -> bytes:
    """Сжимает изображение до max_width px, конвертирует в WebP и возвращает bytes."""
    try:
        img = Image.open(filepath)
        img = img.convert("RGB")
        w, h = img.size
        if w > max_width:
            new_h = int(h * max_width / w)
            img = img.resize((max_width, new_h), Image.LANCZOS)
        buf = io.BytesIO()
        img.save(buf, format="WEBP", quality=quality, method=6)
        return buf.getvalue()
    except Exception:
        return filepath.read_bytes()


def upload_renders_to_s3(jk_name: str) -> int:
    """Сжимает и загружает все файлы из renders/{jk_name}/ в S3.
    Полный размер (1200px) → terrace/renders/{jk}/
    Превью (400px)         → terrace/renders-thumb/{jk}/
    """
    folder = RENDERS_DIR / safe_dirname(jk_name)
    if not folder.exists():
        return 0
    files = sorted(folder.glob("*"))
    if not files:
        return 0
    try:
        s3 = get_s3_client()
        uploaded = 0
        for f in files:
            # Всегда сохраняем как .jpg
            fname_webp = f.stem + '.webp'
            # Полноразмерное (1200px)
            jk_safe = jk_name.replace('/', '')
            key_full  = f"{S3_PREFIX}/{jk_safe}/{fname_webp}"
            data_full = compress_image(f, max_width=1200, quality=85)
            s3.put_object(Bucket=S3_BUCKET, Key=key_full, Body=data_full,
                          ContentType="image/webp", ACL="public-read")
            # Превью (400px)
            key_thumb  = f"terrace/renders-thumb/{jk_safe}/{fname_webp}"
            data_thumb = compress_image(f, max_width=400, quality=75)
            s3.put_object(Bucket=S3_BUCKET, Key=key_thumb, Body=data_thumb,
                          ContentType="image/webp", ACL="public-read")
            uploaded += 1
        return uploaded
    except (BotoCoreError, ClientError) as e:
        print(f"  [!] Ошибка загрузки в S3 ({jk_name}): {e}")
        return 0


def upload_layout_to_s3(filepath: Path, jk: str) -> bool:
    """Загружает планировку в S3 как WebP: terrace/layouts/{jk}/{filename}.webp
    Конвертирует PNG/JPG → WebP — генератор ищет именно .webp"""
    try:
        s3 = get_s3_client()
        jk_safe = safe_dirname(jk)
        webp_name = filepath.stem + ".webp"
        key = f"terrace/layouts/{jk_safe}/{webp_name}"
        img = Image.open(filepath).convert("RGB")
        buf = io.BytesIO()
        img.save(buf, format="WEBP", quality=88, method=6)
        buf.seek(0)
        webp_data = buf.read()
        s3.put_object(
            Bucket=S3_BUCKET, Key=key, Body=webp_data,
            ContentType="image/webp", ACL="public-read",
        )
        # Сохраняем WebP локально — чтобы layout_exists() нашёл его при следующем запуске
        webp_local = filepath.parent / webp_name
        webp_local.write_bytes(webp_data)
        return True
    except (BotoCoreError, ClientError) as e:
        print(f"  [!] Ошибка загрузки планировки в S3: {e}")
        return False
    except Exception as e:
        print(f"  [!] Ошибка конвертации планировки: {e}")
        return False


# ──────────────────────────────────────────────
# БРАУЗЕР
# ──────────────────────────────────────────────
def get_driver():
    CHROME_PROFILE.mkdir(exist_ok=True)
    opts = Options()
    opts.add_argument(f"--user-data-dir={CHROME_PROFILE}")
    opts.add_argument("--profile-directory=Default")
    opts.add_argument("--start-maximized")
    opts.add_argument("--disable-blink-features=AutomationControlled")
    opts.add_experimental_option("excludeSwitches", ["enable-automation"])
    opts.add_experimental_option("useAutomationExtension", False)
    driver = webdriver.Chrome(options=opts)
    driver.execute_cdp_cmd(
        "Page.addScriptToEvaluateOnNewDocument",
        {"source": "Object.defineProperty(navigator,'webdriver',{get:()=>undefined})"}
    )
    return driver


def wait_for_page_ready(driver, timeout=15):
    for _ in range(timeout):
        if driver.execute_script("return document.readyState") == "complete":
            break
        time.sleep(1)
    prev_url = ""
    for _ in range(10):
        cur_url = driver.current_url
        if cur_url == prev_url:
            break
        prev_url = cur_url
        time.sleep(1)


def is_on_sso(driver):
    return "sso.trend.tech" in driver.current_url or "login" in driver.current_url


def ensure_auth(driver):
    print(f"[•] Профиль: {CHROME_PROFILE}")
    for attempt in range(5):
        driver.get(BASE_URL)
        wait_for_page_ready(driver)
        time.sleep(5)  # даём SPA время восстановить своё состояние
        cur = driver.current_url
        if is_on_sso(driver):
            break  # нужна авторизация — выходим
        # Проверяем что наши фильтры в URL (TrendAgent мог перезаписать URL)
        if "apartments-balcony" in cur and "apartments-level_type" in cur:
            break  # фильтры на месте
        print(f"[•] Попытка {attempt+1}: TrendAgent перезаписал URL, переходим заново…")
        time.sleep(2)
    print(f"[•] URL: {driver.current_url[:80]}")
    if not is_on_sso(driver):
        print("[✓] Сессия активна")
        return
    print("[!] Нужен вход — авторизуйтесь в браузере.")
    input("   После входа нажмите Enter …\n")
    print("[•] Ждём завершения авторизации…", end=" ", flush=True)
    for _ in range(30):
        if not is_on_sso(driver):
            break
        time.sleep(1)
    print(f"URL: {driver.current_url[:80]}")
    time.sleep(8)
    print("готово")
    # После авторизации открываем нужный URL с фильтрами
    print("[•] Открываем страницу с фильтрами…")
    _loaded = False
    for _attempt in range(5):
        driver.get(BASE_URL)
        wait_for_page_ready(driver)
        time.sleep(5)  # даём SPA время восстановить состояние
        # Проверяем что URL не был перезаписан TrendAgent-ом
        _cur = driver.current_url
        if "apartments-balcony" in _cur and "apartments-level_type" in _cur:
            _loaded = True
            break
        print(f"[!] Попытка {_attempt+1}: TrendAgent перезаписал URL ({_cur[:60]}…), перезагружаем…")
        time.sleep(2)
    if not _loaded:
        print("[!] Не удалось загрузить страницу с нужными фильтрами — проверьте URL вручную")


# ──────────────────────────────────────────────
# ВСПОМОГАТЕЛЬНЫЕ ФУНКЦИИ
# ──────────────────────────────────────────────
def safe_dirname(name: str) -> str:
    """Убирает символы недопустимые в именах папок."""
    return re.sub(r'[\\/*?:"<>|]', "", name).strip()


def safe_filename(name: str) -> str:
    """Убирает символы недопустимые в именах файлов включая слэш."""
    return re.sub(r'[\\/*?:"<>|/]', "_", name).strip()


def layout_exists(jk: str, apt_id: str) -> bool:
    """Возвращает True только если уже есть .webp.
    Если есть .png/.jpg — конвертирует и заливает в S3, затем возвращает True."""
    folder = LAYOUTS_DIR / safe_dirname(jk)
    fname  = f"{safe_filename(jk)}_{safe_filename(apt_id)}"
    # Уже есть webp — ничего делать не нужно
    if (folder / f"{fname}.webp").exists():
        return True
    # Есть png/jpg — конвертируем в webp и заливаем, затем удаляем исходник
    for ext in ("jpg", "jpeg", "png"):
        f_path = folder / f"{fname}.{ext}"
        if f_path.exists():
            print(f"  [↻] Конвертирую в WebP: {f_path.name}")
            upload_layout_to_s3(f_path, jk)
            try: f_path.unlink()  # удаляем PNG — теперь есть WebP
            except Exception: pass
            return True
    return False


# ──────────────────────────────────────────────
# ТАБЛИЦА + ПЛАНИРОВКИ (за один проход)
# ──────────────────────────────────────────────
def count_rows(driver):
    soup = BeautifulSoup(driver.page_source, "html.parser")
    return len([r for r in soup.find_all("tr") if r.find("td")])


def expand_all(driver):
    print("\n[•] Разворачиваем все строки таблицы…")
    for _ in range(20):
        if count_rows(driver) > 0:
            break
        time.sleep(1)
    time.sleep(3)
    click_num = 0
    while True:
        driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
        time.sleep(2)
        btn = None
        for el in driver.find_elements(By.CSS_SELECTOR, "button"):
            try:
                if el.is_displayed() and ("Показать ещё" in el.text or "Показать еще" in el.text):
                    btn = el
                    break
            except Exception:
                continue
        if not btn:
            break
        prev = count_rows(driver)
        print(f"  [+] Клик {click_num+1} (строк: {prev})…", end=" ", flush=True)
        driver.execute_script("arguments[0].scrollIntoView({block:'center'});", btn)
        time.sleep(0.5)
        try:
            btn.click()
        except Exception:
            driver.execute_script("arguments[0].click();", btn)
        for _ in range(15):
            if count_rows(driver) > prev:
                break
            time.sleep(1)
        new_count = count_rows(driver)
        print(f"+{new_count - prev} (итого {new_count})")
        click_num += 1
        driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
        time.sleep(2)
    print(f"[✓] Строк загружено: {count_rows(driver)} шт.")


def parse_table(soup):
    table = soup.find("table")
    if not table:
        return [], []
    headers = []
    thead = table.find("thead")
    header_row = (thead or table).find("tr")
    if header_row:
        for cell in header_row.find_all(["th", "td"]):
            headers.append(re.sub(r"\s{2,}", " ", cell.get_text(separator=" ", strip=True)))
    print(f"  [✓] Колонок: {len(headers)}  →  {headers}")
    tbody = table.find("tbody")
    rows = []
    for tr in (tbody or table).find_all("tr"):
        tds = tr.find_all("td")
        if not tds:
            continue
        cells = [re.sub(r"\s{2,}", " ", td.get_text(separator=" ", strip=True)) for td in tds]
        rows.append(cells)
    return headers, rows


def download_layout_via_click(driver, row_el, apt_id: str, jk: str) -> bool:
    """Кликает на строку таблицы → новая вкладка → скачивает планировку → закрывает вкладку."""
    folder = LAYOUTS_DIR / safe_dirname(jk)
    folder.mkdir(parents=True, exist_ok=True)
    fname = f"{safe_filename(jk)}_{safe_filename(apt_id)}"

    tabs_before = set(driver.window_handles)
    try:
        driver.execute_script("arguments[0].click();", row_el)
    except Exception:
        return False

    # Ждём новую вкладку
    for _ in range(10):
        time.sleep(0.5)
        new_tabs = set(driver.window_handles) - tabs_before
        if new_tabs:
            break
    else:
        return False

    driver.switch_to.window(new_tabs.pop())
    wait_for_page_ready(driver)
    time.sleep(2)

    # Ждём планировку квартиры — первое изображение из первого слайда
    # apartment-first_slide содержит именно планировку квартиры (не поэтажный план)
    src = ''
    for _wait_i in range(15):
        try:
            src = driver.execute_script("""
                // Берём первый слайд (.apartment-first_slide) — там планировка квартиры
                // Поэтажный план находится в другом контейнере (без класса apartment-first_slide)
                var firstSlide = document.querySelector('.apartment-image_container.apartment-first_slide');
                if (firstSlide) {
                    var imgs = firstSlide.querySelectorAll('img');
                    for (var i = 0; i < imgs.length; i++) {
                        var s = imgs[i].src || imgs[i].getAttribute('data-src') || '';
                        if (s && !s.startsWith('data:') && s.length > 10) return s;
                    }
                }
                // Fallback — первый контейнер не в leaflet
                var containers = document.querySelectorAll('.apartment-image_container');
                for (var ci = 0; ci < containers.length; ci++) {
                    var c = containers[ci];
                    if (c.closest('.leaflet-pane') || c.closest('.leaflet-container')) continue;
                    if (c.closest('.gallery-nav') || c.classList.contains('gallery-nav__item-inner')) continue;
                    var imgs2 = c.querySelectorAll('img');
                    for (var i = 0; i < imgs2.length; i++) {
                        var s2 = imgs2[i].src || imgs2[i].getAttribute('data-src') || '';
                        if (s2 && !s2.startsWith('data:') && s2.length > 10) return s2;
                    }
                }
                return '';
            """) or ''
        except Exception:
            src = ''
        # Считаем URL валидным если он не пустой и не data:
        if src and not src.startswith('data:') and len(src) > 10:
            break
        if _wait_i < 14:
            time.sleep(1)

    ok = False
    if src:
        ext = src.split(".")[-1].split("?")[0].lower()
        if ext not in ("jpg", "jpeg", "png", "webp"):
            ext = "jpg"
        try:
            cookies = {c["name"]: c["value"] for c in driver.get_cookies()}
            resp = requests.get(src, cookies=cookies,
                                headers={"Referer": driver.current_url}, timeout=15)
            if resp.status_code == 200:
                filepath = folder / f"{fname}.{ext}"
                filepath.write_bytes(resp.content)
                # Копируем в папку новых планировок
                new_folder = LAYOUTS_DIR / "_new"
                new_folder.mkdir(exist_ok=True)
                shutil.copy2(filepath, new_folder / f"{fname}.{ext}")
                # Загружаем в S3
                upload_layout_to_s3(filepath, jk)
                ok = True
        except Exception:
            pass

    # Парсим дополнительные данные квартиры
    apt_extra = {}
    try:
        rows_data = driver.execute_script(
            "var rows=document.querySelectorAll('.apartment-passport__row');"
            "var result={};"
            "rows.forEach(function(row){"
            "var text=row.innerText.trim();"
            "var parts=text.split('\\n');"
            "if(parts.length>=2)result[parts[0].trim()]=parts.slice(1).join(' ').trim();"
            "});"
            "return result;"
        )
        if rows_data:
            FIELDS = {
                'Цена за м² при 100% оплате': 'цена_за_м2',
                'Базовая цена':               'базовая_цена',
                'Цена за м² при базовой цене':'цена_за_м2_базовая',
                'Окна':                       'окна',
                'Вид из окна':                'вид_из_окна',
                'Видовая квартира':            'видовая_квартира',
                'Срок сдачи':                 'срок_сдачи',
                'Дата старта продаж':         'дата_старта_продаж',
                'Класс недвижимости':         'класс_недвижимости',
                'Высота потолков':            'высота_потолков',
                'Старт продаж, цена':         'старт_продаж_цена',
            }
            for ru_name, key in FIELDS.items():
                if ru_name in rows_data:
                    apt_extra[key] = rows_data[ru_name]
            # Дебаг: печатаем все строки паспорта квартиры
            if not apt_extra.get('базовая_цена'):
                print(f"  [DEBUG passport rows]: {list(rows_data.keys())}")
            else:
                bp = apt_extra.get('базовая_цена', '')
                p100 = apt_extra.get('цена_за_м2', '')
                print(f"  [DEBUG] Базовая цена: {bp!r}  |  Цена за м² 100%: {p100!r}")
    except Exception as e:
        print(f"  [DEBUG] passport exception: {e}")

    # Закрываем вкладку квартиры, возвращаемся к таблице
    driver.close()
    driver.switch_to.window(driver.window_handles[0])
    time.sleep(1)
    return ok, apt_extra


def parse_and_download(driver, need_layouts: bool) -> tuple:
    """
    Парсит таблицу и скачивает планировки за один проход по строкам.
    Возвращает (headers, rows, layout_ids).
    """
    print("\n[•] Парсим таблицу…")
    soup = BeautifulSoup(driver.page_source, "html.parser")
    headers, rows = parse_table(soup)
    print(f"[✓] Всего строк: {len(rows)}")

    if not rows:
        return headers, rows, set()

    id_idx = headers.index("id / №") if "id / №" in headers else None
    jk_idx = headers.index("ЖК") if "ЖК" in headers else None

    layout_ids = set()
    apt_extra_map = {}  # apt_id -> доп данные

    if not need_layouts:
        return headers, rows, layout_ids, apt_extra_map

    # Получаем элементы строк из DOM
    row_els = driver.find_elements(By.CSS_SELECTOR, "tr.apartments-table__row")
    print(f"\n[•] Скачиваем планировки ({len(rows)} кв.)…")

    for i, cells in enumerate(rows):
        apt_id = cells[id_idx].strip() if id_idx is not None and id_idx < len(cells) else ""
        jk     = cells[jk_idx].strip() if jk_idx is not None and jk_idx < len(cells) else ""

        if not apt_id or not jk:
            continue
        if layout_exists(jk, apt_id):
            layout_ids.add(apt_id)
            continue

        print(f"  [{i+1}/{len(rows)}] {jk} / {apt_id}…", end=" ", flush=True)

        # row_els могут обновляться после возврата — перезапрашиваем
        row_els = driver.find_elements(By.CSS_SELECTOR, "tr.apartments-table__row")
        if i >= len(row_els):
            print("строка не найдена")
            continue

        ok, apt_extra = download_layout_via_click(driver, row_els[i], apt_id, jk)
        print("✓" if ok else "нет планировки")
        if ok:
            layout_ids.add(apt_id)
        # Сохраняем доп данные квартиры
        if apt_extra and apt_id:
            apt_extra_map[f"{jk}_{apt_id}"] = apt_extra

    return headers, rows, layout_ids, apt_extra_map


# ──────────────────────────────────────────────
# СРАВНЕНИЕ С ПРЕДЫДУЩИМ EXCEL
# ──────────────────────────────────────────────
def get_prev_excel() -> Path | None:
    files = sorted(SCRIPT_DIR.glob("trendagent_*.xlsx"), reverse=True)
    return files[0] if files else None


def read_excel_data(path: Path) -> tuple:
    wb = load_workbook(path)
    ws = wb.active
    headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    jk_idx    = headers.index("ЖК") if "ЖК" in headers else None
    id_idx    = headers.index("id / №") if "id / №" in headers else None
    price_idx = next((i for i, h in enumerate(headers) if h and "100%" in str(h)), None)
    jk_set = set()
    prices = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        jk = str(row[jk_idx]).strip() if jk_idx is not None and row[jk_idx] else ""
        if jk:
            jk_set.add(jk)
        if id_idx is not None and row[id_idx]:
            apt_id = str(row[id_idx]).strip()
            price  = str(row[price_idx]).strip() if price_idx is not None and row[price_idx] else ""
            key = f"{jk}__{apt_id}" if jk else apt_id
            prices[key] = price
    return jk_set, prices


def extract_current_data(headers, rows):
    jk_idx    = headers.index("ЖК") if "ЖК" in headers else None
    id_idx    = headers.index("id / №") if "id / №" in headers else None
    price_idx = next((i for i, h in enumerate(headers) if h and "100%" in h), None)
    jk_set = set()
    prices = {}
    for cells in rows:
        jk = cells[jk_idx].strip() if jk_idx is not None and jk_idx < len(cells) and cells[jk_idx] else ""
        if jk:
            jk_set.add(jk)
        if id_idx is not None and id_idx < len(cells) and cells[id_idx]:
            apt_id = cells[id_idx].strip()
            price  = cells[price_idx].strip() if price_idx is not None and price_idx < len(cells) else ""
            key = f"{jk}__{apt_id}" if jk else apt_id
            prices[key] = price
    return jk_set, prices


# ──────────────────────────────────────────────
# РЕНДЕРЫ ЖК
# ──────────────────────────────────────────────
TRANSLIT = {
    'а':'a','б':'b','в':'v','г':'g','д':'d','е':'e','ё':'yo','ж':'zh',
    'з':'z','и':'i','й':'y','к':'k','л':'l','м':'m','н':'n','о':'o',
    'п':'p','р':'r','с':'s','т':'t','у':'u','ф':'f','х':'kh','ц':'ts',
    'ч':'ch','ш':'sh','щ':'sch','ъ':'','ы':'y','ь':'','э':'e','ю':'yu','я':'ya'
}

def translit(s: str) -> str:
    return ''.join(TRANSLIT.get(c, c) for c in s.lower())


def get_object_links(driver) -> list:
    driver.get(LIST_URL)
    wait_for_page_ready(driver)
    time.sleep(3)
    while True:
        driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
        time.sleep(2)
        btn = None
        for el in driver.find_elements(By.CSS_SELECTOR, "button"):
            try:
                if el.is_displayed() and ("Показать ещё" in el.text or "Показать еще" in el.text):
                    btn = el
                    break
            except Exception:
                continue
        if not btn:
            break
        driver.execute_script("arguments[0].scrollIntoView({block:'center'});", btn)
        time.sleep(0.5)
        try:
            btn.click()
        except Exception:
            driver.execute_script("arguments[0].click();", btn)
        time.sleep(3)
    time.sleep(4)
    data = driver.execute_script(
        "var result = [];"
        "document.querySelectorAll('a').forEach(function(a) {"
        "  var href = a.href || '';"
        "  if (href.includes('/object/') && !href.includes('/flat/')) {"
        "    var caption = a.querySelector('.apartment-horizontal-caption');"
        "    var name = caption ? caption.innerText.trim() : '';"
        "    if (!name) {"
        "      var lines = (a.innerText || '').split(String.fromCharCode(10)).map(function(s){return s.trim();}).filter(function(s){return s;});"
        "      name = lines[0] || '';"
        "    }"
        "    result.push({href: href, name: name});"
        "  }"
        "});"
        "return result;"
    )
    result = []
    seen = set()
    for item in data:
        base = item['href'].split("?")[0].rstrip("/")
        if base in seen:
            continue
        seen.add(base)
        slug = base.rstrip("/").split("/")[-1]
        name = item['name'] or slug
        result.append((name, base, slug))
    return result


def find_jk_link(jk_name: str, all_links: list):
    jk_lower = jk_name.lower().strip()
    jk_tr    = translit(jk_lower)

    # 1. Точное совпадение по названию
    for name, url, slug in all_links:
        if name.lower().strip() == jk_lower:
            return name, url

    # 2. Название содержит или содержится
    for name, url, slug in all_links:
        n = name.lower().strip()
        if jk_lower in n or n in jk_lower:
            return name, url

    # 3. Транслит совпадает со slug
    for name, url, slug in all_links:
        if jk_tr == slug.lower() or jk_tr in slug.lower() or slug.lower() in jk_tr:
            return name, url

    # 4. Совпадение по словам
    words    = [w for w in re.split(r'[\s\-_/,.]', jk_lower) if len(w) > 3]
    words_tr = [translit(w) for w in words]
    for name, url, slug in all_links:
        n = name.lower() + " " + slug.lower()
        m1 = sum(1 for w in words if w in n)
        m2 = sum(1 for w in words_tr if w in slug.lower())
        if words and (m1 >= max(1, len(words) - 1) or m2 >= max(1, len(words_tr) - 1)):
            return name, url

    return None, None


def _contains_trend(text: str) -> bool:
    """Проверяет содержит ли текст слово trend (отдельно или в составе)."""
    return 'trend' in text.lower()


def get_fresh_auth_token(driver: "webdriver.Chrome") -> str:
    """Всегда берёт свежий auth_token из кук браузера."""
    try:
        cookies = driver.get_cookies()
        return next((c["value"] for c in cookies if c["name"] == "auth_token"), "")
    except Exception:
        return ""


def parse_jk_data(driver, jk_name: str, url: str, block_id_hint: str = "") -> dict:
    """
    Парсит страницу ЖК — текст, карточки, акции, ипотека, рассрочки.
    """
    import json as _json
    import re as _re
    import urllib.parse as up

    # Скроллим страницу полностью для подгрузки всех блоков
    # Несколько проходов с паузами — ждём lazy-load блоков ипотеки/рассрочек
    for pct in [0.2, 0.5, 0.8, 1.0]:
        driver.execute_script(f"window.scrollTo(0, document.body.scrollHeight * {pct});")
        time.sleep(0.8)
    # Явно скроллим до блоков ипотеки и рассрочек чтобы спровоцировать их загрузку
    driver.execute_script("""
        var selectors = [
            '.object-mortgage', '[class*="mortgage"]', '[class*="installment"]',
            '[class*="рассрочк"]', '[class*="ипотек"]', '[class*="finance"]'
        ];
        for (var i = 0; i < selectors.length; i++) {
            var el = document.querySelector(selectors[i]);
            if (el) { el.scrollIntoView(); break; }
        }
    """)
    time.sleep(2)
    driver.execute_script("window.scrollTo(0, 0);")
    time.sleep(1)

    result = {"about": "", "features": [], "promo": [], "mortgage": [], "installments": []}

    # ── Берём auth_token из куков ─────────────────────────────
    cookies_list = driver.get_cookies()
    auth_token = next((c["value"] for c in cookies_list if c["name"] == "auth_token"), "")

    # ── block_id из performance ────────────────────────────────
    block_id = ""
    city_id  = "58c665588b6aa52311afa01b"

    # Сначала пробуем найти в уже загруженных запросах
    for _ in range(12):
        block_id = driver.execute_script("""
            var entries = performance.getEntriesByType('resource');
            for (var i = 0; i < entries.length; i++) {
                var n = entries[i].name;
                if (n.includes('trendagent.ru')) {
                    var m = n.match(/blocks\\/([a-f0-9]{24})/);
                    if (m) return m[1];
                }
            }
            return '';
        """)
        if block_id:
            break
        time.sleep(1.0)

    # Если не нашли — скроллим до блока ипотеки чтобы спровоцировать загрузку
    if not block_id:
        driver.execute_script("""
            var selectors = [
                '.object-mortgage', '[class*="mortgage"]', '[class*="installment"]',
                '[class*="рассрочк"]', '[class*="ипотек"]'
            ];
            for (var i = 0; i < selectors.length; i++) {
                var el = document.querySelector(selectors[i]);
                if (el) { el.scrollIntoView(); break; }
            }
        """)
        time.sleep(4)
        for _ in range(15):
            block_id = driver.execute_script("""
                var entries = performance.getEntriesByType('resource');
                for (var i = 0; i < entries.length; i++) {
                    var n = entries[i].name;
                    if (n.includes('trendagent.ru')) {
                        var m = n.match(/blocks\\/([a-f0-9]{24})/);
                        if (m) return m[1];
                    }
                }
                return '';
            """)
            if block_id:
                break
            time.sleep(0.5)

    # Последний вариант — берём block_id из URL страницы
    if not block_id:
        try:
            import re as _re
            cur = driver.current_url
            m = _re.search(r'_id%5D=([a-f0-9]{24})', cur)
            if not m:
                m = _re.search(r'\[_id\]=([a-f0-9]{24})', cur)
            if m:
                block_id = m.group(1)
        except:
            pass

    # ── Текст "Об объекте" ────────────────────────────────────
    try:
        driver.execute_script("""
            var el = document.querySelector('.object-about');
            if (el) el.scrollIntoView();
        """)
        time.sleep(1)
        about = driver.execute_script("""
            var els = document.querySelectorAll('.object-about p');
            if (els.length) {
                return Array.from(els).map(e => e.innerText.trim()).filter(t => t).join('\\n\\n');
            }
            var el = document.querySelector('.object-about__text, .object-about p');
            return el ? el.innerText.trim() : '';
        """)
        result["about"] = about or ""
    except Exception:
        pass

    # ── Карточки особенностей ─────────────────────────────────
    try:
        features = driver.execute_script("""
            var items = document.querySelectorAll('.object-about__item, .object-about [class*="item"]');
            var result = [];
            items.forEach(function(el) {
                var text = el.querySelector('.advantages-list__text');
                var img  = el.querySelector('img');
                if (text || img) {
                    result.push({
                        text: text ? text.innerText.trim() : '',
                        img:  img  ? img.src : ''
                    });
                }
            });
            return result;
        """)
        result["features"] = features or []
    except Exception:
        pass

    # ── Акции/скидки ──────────────────────────────────────────
    try:
        # Скроллим до блока скидок чтобы спровоцировать загрузку
        driver.execute_script("""
            var sel = '[class*="discount"], [class*="promo"], [class*="акци"], [class*="скидк"]';
            var el = document.querySelector(sel);
            if (el) el.scrollIntoView();
        """)
        time.sleep(2)
        # Акции: делаем запрос через requests с куками браузера
        _city_d = "58c665588b6aa52311afa01b"
        _builder_d = "58c665588b6aa52311afa0c3"
        try:
            _auth_d = get_fresh_auth_token(driver)  # обновляем токен
            _disc_url = f"https://discounts.trendagent.ru/blocks/{block_id}/discounts?builder={_builder_d}&city={_city_d}&lang=ru"
            if _auth_d:
                _disc_url += f"&auth_token={_auth_d}"
            _disc_resp = requests.get(_disc_url,
                                      headers={"Referer": url}, timeout=10)
            if _disc_resp.status_code == 200:
                _disc_data = _json.loads(_disc_resp.text)
                discounts_list = _disc_data.get("discounts", [])
                # === ДЕБАГ: сохранить сырой JSON первого ЖК в файл ===
                import os as _os_d
                _debug_path = _os_d.path.join(_os_d.path.dirname(_os_d.path.abspath(__file__)), "debug_promo_raw.json")
                if discounts_list and not _os_d.path.exists(_debug_path):
                    with open(_debug_path, "w", encoding="utf-8") as _df:
                        _json.dump(discounts_list[:3], _df, ensure_ascii=False, indent=2)
                    print(f"  [debug] Сырой JSON акций сохранён: {_debug_path}")
                # === КОНЕЦ ДЕБАГА ===
                result["promo"] = [
                    {
                        "name":        d.get("name", ""),
                        "value":       d.get("value", "") or "",
                        "duration":    d.get("duration", [""])[0] if d.get("duration") else "",
                        "description": _re.sub(r'<[^>]+>', '', d.get("description", "") or "").strip(),
                        "conditions":  d.get("promotion_conditions", {}).get("name", "") if isinstance(d.get("promotion_conditions"), dict) else str(d.get("promotion_conditions", "")),
                        "mortgage_combinations": _re.sub(r'<[^>]+>', '', d.get("mortgage_combinations", "") or "").strip(),
                        "installment_combinations": _re.sub(r'<[^>]+>', '', d.get("installment_combinations", "") or "").strip(),
                        "summation": _re.sub(r'<[^>]+>', '', d.get("summation_with_other_discount", "") or "").strip(),
                        "label":       d.get("label", "") or "",
                        "apartments_count": d.get("apartments_count", 0),
                        "required_documents": _re.sub(r'<[^>]+>', '', d.get("required_documents", "") or "").strip(),
                    }
                    for d in discounts_list
                    if d.get("is_active", True) and not _contains_trend(d.get("name", ""))
                ]
                print(f"  Акции: {len(result['promo'])} шт.")
                # Дебаг: показать все акции из API
                print(f"  [debug] Всего в API: {len(discounts_list)}, после фильтра: {len(result['promo'])}")
                for _di, _dd in enumerate(discounts_list):
                    _act = '✓' if _dd.get('is_active', True) else '✗'
                    _trn = ' [TREND]' if _contains_trend(_dd.get('name','')) else ''
                    print(f"    {_act} {_dd.get('name','?')} | {_dd.get('value','')}{_trn}")
                if len(discounts_list) != len(result['promo']):
                    print(f"  [!] Отфильтровано: {len(discounts_list) - len(result['promo'])} акций")
            else:
                print(f"  [!] Акции API: {_disc_resp.status_code}")
        except Exception as _disc_e:
            print(f"  [!] Акции: {_disc_e}")
        if True:
            pass
    except Exception as e:
        print(f"  [!] Акции: {e}")

    print(f"  block_id={block_id!r}")
    if not block_id:
        # Попробуем вытащить из URL страницы
        try:
            cur = driver.current_url
            import re as _re
            m = _re.search(r'_id%5D=([a-f0-9]{24})', cur)
            if not m:
                m = _re.search(r'\[_id\]=([a-f0-9]{24})', cur)
            if m:
                block_id = m.group(1)
                print(f"  block_id из URL: {block_id!r}")
        except:
            pass
    if not block_id and block_id_hint:
        block_id = block_id_hint
        print(f"  block_id из кэша квартиры: {block_id!r}")

    if not block_id:
        print(f"  [!] block_id не найден — ипотека и рассрочки недоступны")
        return result

    # Обновляем auth_token перед API запросами (токен живёт 5 минут)
    cookies_list = driver.get_cookies()
    auth_token = next((c["value"] for c in cookies_list if c["name"] == "auth_token"), auth_token)

    # ── Рассрочки ─────────────────────────────────────────────
    try:
        auth_token = get_fresh_auth_token(driver)  # обновляем токен
        inst_url = f"https://tiny-installments-api.trendagent.ru/v1/blocks/{block_id}?city={city_id}&lang=ru"
        headers = {"Authorization": f"Bearer {auth_token}"} if auth_token else {}
        resp = requests.get(inst_url, headers=headers, timeout=10)
        if resp.status_code == 200:
            data = _json.loads(resp.text)
            if isinstance(data, list):
                import re as _re_inst
                def _strip_html(s):
                    return _re_inst.sub(r'<[^>]+>', '', s or '').strip()
                def _html_to_list(s):
                    items = _re_inst.findall(r'<li[^>]*>(.*?)</li>', s or '', _re_inst.DOTALL)
                    return [_strip_html(i).strip() for i in items if _strip_html(i).strip()]
                result["installments"] = [
                    {
                        "name":                    d.get("name", ""),
                        "firstpay":                d.get("firstpay", ""),
                        "term":                    d.get("term", ""),
                        "price":                   d.get("price", ""),
                        "price_comment":           d.get("price_comment", ""),
                        "term_comment":            d.get("term_comment", ""),
                        "transition_to_mortgage":  d.get("transition_to_mortgage", ""),
                        "maternal_capital":        d.get("maternal_capital", ""),
                        "comprehensive_insurance": d.get("comprehensive_insurance", ""),
                        "keys_before_full_payment":d.get("keys_before_full_payment", ""),
                        "receiving_key_conditions":d.get("receiving_key_conditions", ""),
                        "comment":                 _html_to_list(d.get("comment", "")),
                        "payment_schedule":        _html_to_list(d.get("payment_schedule", "")),
                        "tags":                    [t.get("name","") for t in d.get("tags", []) if t.get("name")],
                    }
                    for d in data
                    if not _contains_trend(d.get("name", ""))
                ]
                print(f"  Рассрочки API: {len(result['installments'])} шт.")
        else:
            print(f"  [!] Рассрочки API: {resp.status_code}")
    except Exception as e:
        print(f"  [!] Рассрочки: {e}")

    # ── Ипотека ───────────────────────────────────────────────
    auth_token = get_fresh_auth_token(driver)  # обновляем токен
    # Типы ипотеки (статичный справочник)
    MORTGAGE_TYPES = {
        "670674312581253574cf1868": "Семейная от застройщика",
        "613a29d58c67b73c38a6b1ce": "Стандартная от застройщика",
        "670674312581253574cf186e": "IT от застройщика",
        "5f917bc73bff70cbf9d3dcc8": "Семейная стандартная",
        "5983801cd07ed144bb7cca33": "Стандартная новостройка",
        "628ba48ab0d40825f40fb41d": "IT стандартная",
        "5f917bb23bff70cbf9d3dcc7": "Сельская",
        "63500eb4fac58e5eb6727f83": "Траншевая",
        "5f917bfe3bff70cbf9d3dcca": "Апартаменты стандартная",
        "613a2a168c67b73c38a6b1cf": "Апартаменты от застройщика",
        "5983801cd07ed144bb7cca32": "Без ПВ от банка",
        "5983801cd07ed144bb7cca31": "Ипотека для ПДКП",
        "670674312581253574cf186a": "Семейная комбо от застройщика",
        "670674312581253574cf186b": "Семейная со сверхлимитом",
        "670674312581253574cf186c": "Военная семейная",
        "5983801cd07ed144bb7cca30": "Военная стандартная",
        "670674312581253574cf186f": "IT со сверхлимитом",
        "670674312581253574cf1870": "IT комбо от застройщика",
        "670674312581253574cf1875": "Траншевая апартаменты",
        "5f917c213bff70cbf9d3dccc": "Коммерция стандартная",
        "670674312581253574cf1876": "Коммерция от застройщика",
        "5983801cd07ed144bb7cca34": "Новостройка акция",
        "5b434fd9d12679ed9c5bb9be": "Готовое. Вторичная",
        "67a35a80a90290436ba4e21b": "Новостройка. Вторичная",
        "6893dee394c64681faf019ca": "Готовая. Вторичная БС",
        "670674312581253574cf1869": "Семейная ИЖС",
        "670674312581253574cf1872": "Стандартная новостройка ИЖС",
        "5983801cd07ed144bb7cca2e": "Загородная. Вторичная",
        "5983801cd07ed144bb7cca2c": "ГС стандартная",
        "5983801cd07ed144bb7cca2d": "ГС от застройщика",
        "5983801cd07ed144bb7cca2f": "Стандартная акция",
        "5f917be93bff70cbf9d3dcc9": "Акция вторичная",
        "613a288d8c67b73c38a6b1c1": "ГС акция",
        "5f917c353bff70cbf9d3dccd": "Кредитование машиномест",
        "5f917c103bff70cbf9d3dccb": "Апартаменты. Вторичная",
    }
    # Программы которые нам интересны (семейная и стандартная)
    FAMILY_TYPES = {
        "670674312581253574cf1868",  # Семейная от застройщика
        "5f917bc73bff70cbf9d3dcc8",  # Семейная стандартная
        "670674312581253574cf186a",  # Семейная комбо от застройщика
        "670674312581253574cf186b",  # Семейная со сверхлимитом
    }
    STD_TYPES = {
        "613a29d58c67b73c38a6b1ce",  # Стандартная от застройщика
        "5983801cd07ed144bb7cca33",  # Стандартная новостройка
        "5f917bfe3bff70cbf9d3dcca",  # Апартаменты стандартная
        "5983801cd07ed144bb7cca34",  # Новостройка акция
    }
    IT_TYPES = {
        "670674312581253574cf186e",  # IT от застройщика
        "628ba48ab0d40825f40fb41d",  # IT стандартная
        "670674312581253574cf186f",  # IT со сверхлимитом
        "670674312581253574cf1870",  # IT комбо от застройщика
    }
    # Застройщик субсидирует — тоже считаем как семейная/стандартная
    BUILDER_SUBSIDY = {
        "613a2a168c67b73c38a6b1cf",  # Апартаменты от застройщика
        "63500eb4fac58e5eb6727f83",  # Траншевая
        "670674312581253574cf1875",  # Траншевая апартаменты
    }

    try:
        mort_url = f"https://mortgage-api.trendagent.ru/blocks/{block_id}/?premiseType=apartment"
        if auth_token:
            mort_url += f"&auth_token={auth_token}&city={city_id}&lang=ru"
        resp = requests.get(mort_url, timeout=10)
        if resp.status_code == 200:
            data = _json.loads(resp.text)
            programs = data.get("data", {}).get("results", [])

            # Собираем все уникальные программы
            seen = set()
            mortgage = []

            for p in programs:
                type_id   = p.get("type", "")
                rate_obj  = p.get("rate", {})
                rate_min  = rate_obj.get("min")
                firstpay  = p.get("firstpay")
                period    = p.get("period")  # срок кредита в годах
                bank_name = p.get("bank", {}).get("name", "") if isinstance(p.get("bank"), dict) else ""

                # Субсидированная ставка — строим подсказку и детальные поля
                rate_note    = ""
                first_rate   = None
                first_months = None
                second_rate  = None
                if rate_obj.get("type") == "variable":
                    fp = rate_obj.get("periods", {}).get("first_period", {})
                    sp = rate_obj.get("periods", {}).get("second_period", {})
                    first_rate   = fp.get("rate_min")
                    first_months = fp.get("quantity_of_months")
                    second_rate  = sp.get("rate_min")
                    if second_rate and first_months:
                        rate_note = f"затем {second_rate}% с {first_months + 1} мес"
                    elif second_rate:
                        rate_note = f"затем {second_rate}%"
                if rate_min is None:
                    continue

                # Берём имя из справочника, если нет — пропускаем неизвестные
                prog_name = MORTGAGE_TYPES.get(type_id, "")
                if not prog_name:
                    continue  # тип не из нашего справочника — пропускаем

                # Дедублируем по имени программы, берём минимальную ставку
                if prog_name in seen:
                    for m in mortgage:
                        if m["prog"] == prog_name:
                            r = float(str(rate_min))
                            if r < float(str(m["rate"]).rstrip("%")):
                                m["rate"]         = f"{rate_min}%"
                                m["rate_note"]    = rate_note
                                m["bank"]         = bank_name
                                m["first_rate"]   = first_rate
                                m["first_months"] = first_months
                                m["second_rate"]  = second_rate
                                m["period"]       = period
                                if firstpay:
                                    pv = str(firstpay).rstrip("0").rstrip(".")
                                    m["pv"] = f"от {pv}%"
                    continue
                if _contains_trend(prog_name):
                    continue
                seen.add(prog_name)

                rate_str = str(rate_min)
                pv_str = str(firstpay).rstrip("0").rstrip(".") if firstpay and "." in str(firstpay) else str(firstpay or "")
                mortgage.append({
                    "prog":         prog_name,
                    "rate":         f"{rate_str}%",
                    "pv":           f"от {pv_str}%" if pv_str else "",
                    "rate_note":    rate_note,
                    "bank":         bank_name,
                    "first_rate":   first_rate,
                    "first_months": first_months,
                    "second_rate":  second_rate,
                    "period":       period,
                    "default":      False,
                })

            # Сортируем по приоритету (семейная первая)
            priority = ["Семейная от застройщика", "Семейная стандартная", "IT от застройщика",
                        "IT стандартная", "Стандартная от застройщика", "Стандартная новостройка",
                        "Апартаменты от застройщика", "Апартаменты стандартная", "Траншевая"]
            mortgage.sort(key=lambda m: priority.index(m["prog"]) if m["prog"] in priority else 99)

            # Если совсем нет данных — заглушки
            if not mortgage:
                mortgage = [
                    {"prog": "Семейная",    "rate": "6%",  "pv": "от 20%", "default": True},
                    {"prog": "Стандартная", "rate": "19%", "pv": "от 20%", "default": True},
                ]

            result["mortgage"] = mortgage
            print(f"  Ипотека API: {[(m['prog'], m['rate'], m['pv']) for m in mortgage]}")
        else:
            print(f"  [!] Ипотека API: {resp.status_code}")
    except Exception as e:
        print(f"  [!] Ипотека: {e}")

    return result


def save_jk_data(jk_name: str, data: dict):
    """Обновляет jk_data.json и загружает в S3."""
    # Читаем существующий файл
    all_data = {}
    if JK_DATA_FILE.exists():
        try:
            all_data = json.loads(JK_DATA_FILE.read_text(encoding="utf-8"))
        except Exception:
            pass
    # Обновляем данные для ЖК
    all_data[jk_name] = data
    # Сохраняем локально
    JK_DATA_FILE.write_text(json.dumps(all_data, ensure_ascii=False, indent=2), encoding="utf-8")
    # Загружаем в S3
    try:
        s3 = get_s3_client()
        s3.put_object(
            Bucket=S3_BUCKET,
            Key=S3_JK_DATA_KEY,
            Body=JK_DATA_FILE.read_bytes(),
            ContentType="application/json",
            ACL="public-read"
        )
        print(f"  [☁] jk_data.json → S3")
    except Exception as e:
        print(f"  [!] Ошибка загрузки jk_data.json: {e}")


def download_renders(driver, name: str, url: str) -> dict:
    """Скачивает рендеры и парсит данные ЖК. Возвращает dict с кол-вом фото и данными ЖК."""
    folder = RENDERS_DIR / safe_dirname(name)
    folder.mkdir(parents=True, exist_ok=True)
    driver.get(url)
    wait_for_page_ready(driver)
    time.sleep(1)
    # Один скролл вниз чтобы lazy-load подгрузил картинки
    driver.execute_script("window.scrollTo(0, document.body.scrollHeight * 0.5);")
    time.sleep(0.5)
    driver.execute_script("window.scrollTo(0, 0);")
    time.sleep(0.3)

    # Скроллим страницу до конца чтобы подгрузить все блоки включая акции
    for pct in [0.3, 0.6, 0.9, 1.0]:
        driver.execute_script(f"window.scrollTo(0, document.body.scrollHeight * {pct});")
        time.sleep(0.5)
    driver.execute_script("window.scrollTo(0, 0);")
    time.sleep(1)
    # Открываем галерею чтобы загрузились все миниатюры
    gallery_opened = False
    try:
        driver.execute_script("""
            var el = document.querySelector('.gallery-nav__item-inner');
            if (el) el.click();
        """)
        from selenium.webdriver.support.ui import WebDriverWait
        from selenium.webdriver.support import expected_conditions as EC
        from selenium.webdriver.common.by import By
        WebDriverWait(driver, 8).until(
            EC.presence_of_element_located((By.CSS_SELECTOR, '.gallery__overlay.popup__overlay_opened'))
        )
        gallery_opened = True
        # Скроллим nav для подгрузки всех миниатюр
        for _ in range(5):
            driver.execute_script("""
                var overlay = document.querySelector('.gallery__overlay.popup__overlay_opened');
                var nav = overlay ? overlay.querySelector('.gallery-nav__menu') : null;
                if (nav) nav.scrollLeft += 9999;
            """)
            time.sleep(0.5)
        # Ждём стабилизации количества
        prev_count = 0
        for _ in range(10):
            time.sleep(0.5)
            count = driver.execute_script("""
                var overlay = document.querySelector('.gallery__overlay.popup__overlay_opened');
                return overlay ? overlay.querySelectorAll('.gallery-nav__item-inner img').length : 0;
            """) or 0
            if count == prev_count and count > 0:
                break
            prev_count = count
    except Exception as e:
        pass
    # Берём все миниатюры из открытой галереи и убираем l_ для полного размера
    srcs = driver.execute_script(r"""
        var seen = {};
        var imgs = [];
        var overlay = document.querySelector('.gallery__overlay.popup__overlay_opened');
        var els = overlay ? overlay.querySelectorAll('.gallery-nav__item-inner img') : document.querySelectorAll('.gallery-nav__item-inner img');
        els.forEach(function(img) {
            var s = img.src || img.getAttribute('data-src') || '';
            if (!s || s.startsWith('data:') || !s.includes('selcdn')) return;
            s = s.replace(/\/l_([^\/]+)$/, '/$1');
            if (!seen[s]) { seen[s] = true; imgs.push(s); }
        });
        return imgs;
    """) or []
    if not srcs:
        return 0
    seen_urls = set()
    seen_files = set()
    full_srcs = []
    for s in srcs:
        s = s.strip()
        if not s:
            continue
        # Дедублируем по полному URL и по имени файла (без query params)
        file_key = s.split("?")[0].split("/")[-1].lower()
        if s in seen_urls or file_key in seen_files:
            continue
        seen_urls.add(s)
        seen_files.add(file_key)
        full_srcs.append(s)
    cookies = {c["name"]: c["value"] for c in driver.get_cookies()}
    downloaded = 0
    for i, src in enumerate(full_srcs, start=1):
        try:
            # Всегда сохраняем как .jpg (локально, конвертация в webp при загрузке в S3)
            filepath = folder / f"{i:03d}.jpg"

            resp = requests.get(src, cookies=cookies, headers={"Referer": url}, timeout=15)
            if resp.status_code == 200:
                filepath.write_bytes(resp.content)
                downloaded += 1
        except Exception:
            pass
    return downloaded


# ──────────────────────────────────────────────
# EXCEL
# ──────────────────────────────────────────────
HEADER_FILL = PatternFill("solid", start_color="1F4E79")
HEADER_FONT = Font(name="Arial", bold=True, color="FFFFFF", size=10)
DATA_FONT   = Font(name="Arial", size=10)
CENTER      = Alignment(horizontal="center", vertical="center")
THIN        = Side(style="thin")
THIN_BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
ALT_FILL    = PatternFill("solid", start_color="DCE6F1")


def build_excel(headers, rows, date_str, layout_ids: set, apt_extra_map: dict = None) -> Path:
    wb = Workbook()
    ws = wb.active
    ws.title = "Квартиры"
    EXTRA_COLS = [
        "Цена за м²", "Базовая цена", "Цена за м² (баз.)",
        "Окна", "Вид из окна", "Видовая квартира", "Срок сдачи (карточка)",
        "Дата старта продаж", "Класс", "Высота потолков", "Старт продаж, цена", "Планировка"
    ]
    all_headers = list(headers) + EXTRA_COLS
    apt_extra_map = apt_extra_map or {}
    n_cols = len(all_headers)

    for col_idx, col_name in enumerate(all_headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTER
        cell.border = THIN_BORDER
    ws.row_dimensions[1].height = 22

    id_idx = headers.index("id / №") if "id / №" in headers else None
    jk_idx = headers.index("ЖК") if "ЖК" in headers else None

    for row_idx, cells in enumerate(rows, start=2):
        alt    = (row_idx % 2 == 0)
        apt_id = cells[id_idx].strip() if id_idx is not None and id_idx < len(cells) else ""
        jk     = cells[jk_idx].strip() if jk_idx is not None and jk_idx < len(cells) else ""
        layout_name = f"{safe_filename(jk)}_{safe_filename(apt_id)}" if apt_id in layout_ids else ""
        extra = apt_extra_map.get(f"{jk}_{apt_id}", apt_extra_map.get(apt_id, {}))
        extra_values = [
            extra.get("цена_за_м2", ""),
            extra.get("базовая_цена", ""),
            extra.get("цена_за_м2_базовая", ""),
            extra.get("окна", ""),
            extra.get("вид_из_окна", ""),
            extra.get("видовая_квартира", ""),
            extra.get("срок_сдачи", ""),
            extra.get("дата_старта_продаж", ""),
            extra.get("класс_недвижимости", ""),
            extra.get("высота_потолков", ""),
            extra.get("старт_продаж_цена", ""),
            layout_name,
        ]
        padded = list(cells)[:len(headers)] + [""] * len(headers)
        padded = padded[:len(headers)] + extra_values

        for col_idx, value in enumerate(padded[:n_cols], start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = DATA_FONT
            cell.border = THIN_BORDER
            cell.alignment = CENTER
            if alt:
                cell.fill = ALT_FILL

    for col_idx in range(1, n_cols + 1):
        max_len = len(str(all_headers[col_idx - 1]))
        for row_idx in range(2, len(rows) + 2):
            val = ws.cell(row=row_idx, column=col_idx).value or ""
            max_len = max(max_len, len(str(val)))
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 3, 40)

    ws.auto_filter.ref = ws.dimensions
    ws.freeze_panes = "A2"
    out_path = SCRIPT_DIR / f"trendagent_{date_str}.xlsx"
    wb.save(out_path)
    return out_path


# ──────────────────────────────────────────────
# MAIN
# ──────────────────────────────────────────────

def reupload_all_layouts():
    """Перезаливает все локальные планировки в S3 как WebP.
    Запускается когда FORCE_REUPLOAD_LAYOUTS = True."""
    print("\n🔄 FORCE_REUPLOAD_LAYOUTS=True — перезаливаем все планировки в WebP...")
    if not LAYOUTS_DIR.exists():
        print("  ⚠️  Папка layouts/ не найдена — нечего заливать")
        return
    all_files = [f for f in LAYOUTS_DIR.rglob("*")
                 if f.suffix.lower() in ('.jpg', '.jpeg', '.png', '.webp')
                 and f.is_file()
                 and f.parent.name != "_new"]
    if not all_files:
        print("  ⚠️  Планировок не найдено в layouts/")
        return
    print(f"  Найдено: {len(all_files)} файлов")
    ok = fail = 0
    for f in all_files:
        jk = f.parent.name
        result = upload_layout_to_s3(f, jk)
        if result:
            ok += 1
            print(f"  ✅ {jk}/{f.stem}.webp")
        else:
            fail += 1
    print(f"\n  Готово: ✅ {ok} залито, ❌ {fail} ошибок")


def main():
    import time as _time
    _start = _time.time()
    date_str = datetime.now().strftime("%Y-%m-%d_%H-%M")
    LAYOUTS_DIR.mkdir(exist_ok=True)
    RENDERS_DIR.mkdir(exist_ok=True)

    driver = get_driver()
    try:
        # 1. Авторизация
        ensure_auth(driver)

        # 2. Предыдущий Excel
        prev_excel = get_prev_excel()
        if prev_excel:
            print(f"[•] Предыдущий Excel: {prev_excel.name}")
            old_jk, old_prices = read_excel_data(prev_excel)
        else:
            print("[•] Первый запуск — предыдущего Excel нет")
            old_jk, old_prices = set(), {}
        old_ids = set(old_prices.keys())

        # 3. Раскрываем таблицу
        # Проверяем что URL содержит правильные фильтры перед парсингом
        _cur_url = driver.current_url
        if "apartments-balcony" not in _cur_url or "apartments-level_type" not in _cur_url:
            print(f"[!] Неверный URL перед парсингом: {_cur_url[:80]}")
            print("[•] Принудительно открываем BASE_URL с фильтрами…")
            for _fix_attempt in range(5):
                driver.get(BASE_URL)
                wait_for_page_ready(driver)
                time.sleep(5)
                _cur_url = driver.current_url
                if "apartments-balcony" in _cur_url and "apartments-level_type" in _cur_url:
                    print(f"[✓] Фильтры восстановлены (попытка {_fix_attempt+1})")
                    break
                print(f"[!] Попытка {_fix_attempt+1} не удалась, повторяем…")
                time.sleep(3)
            else:
                print("[!] КРИТИЧНО: не удалось восстановить фильтры — завершение")
                return
        expand_all(driver)

        # 4. Парсим таблицу + скачиваем планировки за один проход
        soup = BeautifulSoup(driver.page_source, "html.parser")
        headers, rows = parse_table(soup)
        print(f"[✓] Всего строк: {len(rows)}")
        if not rows:
            print("[!] Данных нет — завершение")
            return


        # Определяем каким квартирам нужна планировка
        new_jk, new_prices = extract_current_data(headers, rows)
        new_ids = set(new_prices.keys())
        added_apts   = new_ids - old_ids
        removed_apts = old_ids - new_ids
        added_jk     = new_jk - old_jk
        changed_prices = {
            apt_id: (old_prices[apt_id], new_prices[apt_id])
            for apt_id in new_ids & old_ids
            if new_prices[apt_id] != old_prices[apt_id]
        }

        # Очищаем папку новых планировок перед запуском
        new_folder = LAYOUTS_DIR / "_new"
        if new_folder.exists():
            shutil.rmtree(new_folder)
        new_folder.mkdir(exist_ok=True)

        first_run = not prev_excel
        id_idx = headers.index("id / №") if "id / №" in headers else None
        jk_idx = headers.index("ЖК") if "ЖК" in headers else None

        # Определяем строки для скачивания планировки
        to_download = []
        layout_ids  = set()
        for i, cells in enumerate(rows):
            apt_id = cells[id_idx].strip() if id_idx is not None and id_idx < len(cells) else ""
            jk     = cells[jk_idx].strip() if jk_idx is not None and jk_idx < len(cells) else ""
            if not apt_id or not jk:
                continue
            if layout_exists(jk, apt_id):
                layout_ids.add(apt_id)  # уже есть
                continue
            # Скачиваем если планировки нет — независимо от того новая квартира или нет
            to_download.append((i, apt_id, jk))

        # Скачиваем планировки
        newly_downloaded = 0
        apt_extra_map = {}
        if to_download:
            print(f"\n[•] Скачиваем планировки ({len(to_download)} кв.)…")
            for n, (i, apt_id, jk) in enumerate(to_download, 1):
                print(f"  [{n}/{len(to_download)}] {jk} / {apt_id}…", end=" ", flush=True)
                row_els = driver.find_elements(By.CSS_SELECTOR, "tr.apartments-table__row")
                if i >= len(row_els):
                    print("строка не найдена")
                    continue
                ok, apt_extra = download_layout_via_click(driver, row_els[i], apt_id, jk)
                print("✓" if ok else "нет планировки")
                if ok:
                    layout_ids.add(apt_id)
                    newly_downloaded += 1
                if apt_extra and apt_id:
                    apt_extra_map[f"{jk}_{apt_id}"] = apt_extra
        else:
            print("\n[✓] Новых планировок нет")

        # 4.5 Собираем extra данные для ВСЕХ квартир (обновляем при каждом запуске)
        # Загружаем кэш из предыдущего запуска (если парсер упал — не начинаем заново)
        EXTRA_CACHE_FILE = SCRIPT_DIR / "apt_extra_cache.json"
        _reset_extra = "--reset-extra" in sys.argv
        if _reset_extra:
            print("[•] --reset-extra: сбрасываем кэш доп. данных, перепарсим все квартиры")
            apt_extra_map = {}
            try: EXTRA_CACHE_FILE.write_text("{}", encoding="utf-8")
            except Exception: pass
        elif EXTRA_CACHE_FILE.exists():
            try:
                cached = json.loads(EXTRA_CACHE_FILE.read_text(encoding="utf-8"))
                apt_extra_map.update(cached)
                print(f"[•] Кэш доп. данных: {len(cached)} кв. загружено")
            except Exception as e:
                print(f"[!] Ошибка чтения кэша: {e}")

        # Пропускаем квартиры у которых уже есть все основные поля
        # При --reset-extra принудительно перепарсиваем все квартиры
        REQUIRED_FIELDS = {'базовая_цена', 'видовая_квартира'}
        apts_without_extra = []
        for i, cells in enumerate(rows):
            if id_idx is None or id_idx >= len(cells): continue
            apt_id = cells[id_idx].strip()
            if not apt_id: continue
            jk_for_key = cells[jk_idx].strip() if jk_idx is not None and jk_idx < len(cells) else ''
            cache_key = f"{jk_for_key}_{apt_id}"
            existing = apt_extra_map.get(cache_key, {})
            # При --reset-extra всегда добавляем в список (кэш уже очищен)
            # Иначе пропускаем если все нужные поля уже есть
            if not _reset_extra and all(f in existing for f in REQUIRED_FIELDS):
                continue
            apts_without_extra.append((i, apt_id, jk_for_key))

        if apts_without_extra:
            print(f"\n[•] Собираем доп. данные квартир ({len(apts_without_extra)} шт., "
                  f"уже есть: {len(rows) - len(apts_without_extra)})…")
            _step_start = time.time()
            for n, (i, apt_id, jk_for_key) in enumerate(apts_without_extra, 1):
                # ETA
                if n > 1:
                    elapsed = time.time() - _step_start
                    per_apt = elapsed / (n - 1)
                    remaining = int(per_apt * (len(apts_without_extra) - n + 1))
                    eta = f"~{remaining//60}м{remaining%60:02d}с" if remaining >= 60 else f"~{remaining}с"
                else:
                    eta = "?"
                print(f"  [{n}/{len(apts_without_extra)} ETA:{eta}] {apt_id}…", end=" ", flush=True)
                try:
                    # Ищем строку по тексту ячейки с apt_id — надёжнее чем по индексу,
                    # т.к. DOM мог обновиться после возврата с предыдущей вкладки
                    row_el = driver.execute_script("""
                        var aptId = arguments[0];
                        var jkName = arguments[1];
                        var rows = document.querySelectorAll('tr.apartments-table__row');
                        // Сначала ищем строку где есть и ЖК и номер квартиры
                        for (var i = 0; i < rows.length; i++) {
                            var rowText = rows[i].textContent;
                            if (rowText.includes(jkName)) {
                                var cells = rows[i].querySelectorAll('td');
                                for (var j = 0; j < cells.length; j++) {
                                    if (cells[j].textContent.trim() === aptId) {
                                        return rows[i];
                                    }
                                }
                            }
                        }
                        // Fallback — только по номеру (если ЖК не в таблице)
                        for (var i = 0; i < rows.length; i++) {
                            var cells = rows[i].querySelectorAll('td');
                            for (var j = 0; j < cells.length; j++) {
                                if (cells[j].textContent.trim() === aptId) {
                                    return rows[i];
                                }
                            }
                        }
                        return null;
                    """, apt_id, jk_for_key)

                    if not row_el:
                        print("строка не найдена — пропускаем")
                        continue

                    # Перехватываем window.open до клика — получаем уникальный ID квартиры из URL
                    expected_flat_id = driver.execute_script("""
                        window._lastOpenedUrl = null;
                        var orig = window.open;
                        window.open = function(url) {
                            window._lastOpenedUrl = url;
                            return orig.apply(this, arguments);
                        };
                        return null;
                    """)

                    tabs_before = set(driver.window_handles)
                    driver.execute_script("arguments[0].click();", row_el)

                    # Ждём открытия вкладки — до 2 повторных попыток
                    new_tabs = set()
                    for attempt in range(3):
                        for _ in range(10):
                            time.sleep(0.5)
                            new_tabs = set(driver.window_handles) - tabs_before
                            if new_tabs:
                                break
                        if new_tabs:
                            break
                        if attempt < 2:
                            print(f"retry{attempt+1}…", end=" ", flush=True)
                            for h in set(driver.window_handles) - tabs_before - {driver.window_handles[0]}:
                                try: driver.switch_to.window(h); driver.close()
                                except: pass
                            driver.switch_to.window(driver.window_handles[0])
                            tabs_before = set(driver.window_handles)
                            time.sleep(1)
                            driver.execute_script("arguments[0].click();", row_el)

                    if not new_tabs:
                        print("вкладка не открылась")
                        continue

                    # Получаем перехваченный URL и извлекаем уникальный ID квартиры
                    captured_url = driver.execute_script("return window._lastOpenedUrl || '';") or ''
                    import re as _re
                    flat_id_match = _re.search(r'/flat/([0-9a-f]{24})', captured_url)
                    expected_flat_id = flat_id_match.group(1) if flat_id_match else ''

                    driver.switch_to.window(new_tabs.pop())
                    wait_for_page_ready(driver)
                    time.sleep(2)

                    # Верификация по уникальному MongoDB ID в URL карточки
                    actual_url = driver.current_url or ''
                    actual_flat_id_match = _re.search(r'/flat/([0-9a-f]{24})', actual_url)
                    actual_flat_id = actual_flat_id_match.group(1) if actual_flat_id_match else ''

                    if expected_flat_id and actual_flat_id and expected_flat_id != actual_flat_id:
                        print(f"⚠️  верификация: ожидался ID {expected_flat_id}, открылся {actual_flat_id} — пропускаем")
                        driver.close()
                        driver.switch_to.window(driver.window_handles[0])
                        time.sleep(1)
                        continue

                    # Ждём появления паспорта
                    for _ in range(10):
                        chk = driver.execute_script(
                            "var r=document.querySelectorAll('.apartment-passport__row');"
                            "return r.length>0?r[0].innerText.trim():'';"
                        ) or ''
                        if chk and '\u200c' not in chk and len(chk) > 3:
                            break
                        time.sleep(0.5)

                    rows_data = driver.execute_script("""
                        var rows = document.querySelectorAll('.apartment-passport__row');
                        var result = {};
                        rows.forEach(function(row) {
                            var text = row.innerText.trim();
                            var parts = text.split('\\n');
                            if (parts.length >= 2) result[parts[0].trim()] = parts.slice(1).join(' ').trim();
                        });
                        return result;
                    """)
                    if rows_data:
                        FIELDS = {
                            'Цена за м² при 100% оплате': 'цена_за_м2',
                            'Базовая цена':               'базовая_цена',
                            'Цена за м² при базовой цене':'цена_за_м2_базовая',
                            'Окна':                       'окна',
                            'Вид из окна':                'вид_из_окна',
                            'Видовая квартира':            'видовая_квартира',
                            'Срок сдачи':                 'срок_сдачи',
                            'Дата старта продаж':         'дата_старта_продаж',
                            'Класс недвижимости':         'класс_недвижимости',
                            'Высота потолков':            'высота_потолков',
                            'Старт продаж, цена':         'старт_продаж_цена',
                        }
                        apt_extra = {key: rows_data[ru] for ru, key in FIELDS.items() if ru in rows_data}
                        cache_key = f"{jk_for_key}_{apt_id}"
                        apt_extra_map[cache_key] = apt_extra
                        # Сохраняем кэш после каждой квартиры — если упадём, не потеряем данные
                        try:
                            EXTRA_CACHE_FILE.write_text(
                                json.dumps(apt_extra_map, ensure_ascii=False, indent=2),
                                encoding="utf-8"
                            )
                        except Exception:
                            pass

                        # Перехватываем block_id из network запросов карточки квартиры
                        # и сохраняем в jk_data.json — шаг 4.6 использует его напрямую
                        if jk_for_key and JK_DATA_FILE.exists():
                            try:
                                _jkd = json.loads(JK_DATA_FILE.read_text(encoding="utf-8"))
                                if jk_for_key in _jkd and not _jkd[jk_for_key].get("block_id"):
                                    _bid = driver.execute_script("""
                                        var e = performance.getEntriesByType('resource');
                                        for (var i=0;i<e.length;i++){
                                            if(e[i].name.includes('trendagent.ru')){
                                                var m=e[i].name.match(/blocks\\/([a-f0-9]{24})/);
                                                if(m) return m[1];
                                            }
                                        }
                                        return '';
                                    """) or ""
                                    if _bid:
                                        _jkd[jk_for_key]["block_id"] = _bid
                                        JK_DATA_FILE.write_text(
                                            json.dumps(_jkd, ensure_ascii=False, indent=2),
                                            encoding="utf-8"
                                        )
                                        print(f" [block_id→{_bid[:8]}…]", end="")
                            except Exception:
                                pass
                    driver.close()
                    driver.switch_to.window(driver.window_handles[0])
                    time.sleep(0.5)
                    print("✓")
                except Exception as e:
                    print(f"[!] {e}")
                    try:
                        driver.close()
                        driver.switch_to.window(driver.window_handles[0])
                    except:
                        pass
        else:
            print("\n[✓] Все доп. данные квартир уже в кэше — пропускаем шаг 4.5")

        # Чистим кэш от квартир которых нет в базе — только если запись старше 30 дней
        # (квартира может быть в броне и вернуться)
        CACHE_META_FILE = SCRIPT_DIR / "apt_extra_cache_meta.json"
        cache_meta = {}
        if CACHE_META_FILE.exists():
            try:
                cache_meta = json.loads(CACHE_META_FILE.read_text(encoding="utf-8"))
            except Exception:
                pass

        now_ts = time.time()
        THIRTY_DAYS = 30 * 24 * 3600

        current_keys = set()
        for cells in rows:
            apt_id = cells[id_idx].strip() if id_idx is not None and id_idx < len(cells) else ''
            jk_k   = cells[jk_idx].strip() if jk_idx is not None and jk_idx < len(cells) else ''
            if apt_id:
                k = f"{jk_k}_{apt_id}"
                current_keys.add(k)
                cache_meta[k] = now_ts  # обновляем метку — квартира активна

        # Удаляем только те которых нет уже 30+ дней
        stale = [k for k in list(apt_extra_map.keys())
                 if k not in current_keys
                 and now_ts - cache_meta.get(k, now_ts) > THIRTY_DAYS]
        if stale:
            for k in stale:
                del apt_extra_map[k]
                cache_meta.pop(k, None)
            try:
                EXTRA_CACHE_FILE.write_text(
                    json.dumps(apt_extra_map, ensure_ascii=False, indent=2), encoding="utf-8"
                )
                CACHE_META_FILE.write_text(
                    json.dumps(cache_meta, ensure_ascii=False, indent=2), encoding="utf-8"
                )
            except Exception:
                pass
            print(f"[•] Кэш: удалено {len(stale)} записей старше 30 дней")
        else:
            # Всё равно обновляем метки активных квартир
            try:
                CACHE_META_FILE.write_text(
                    json.dumps(cache_meta, ensure_ascii=False, indent=2), encoding="utf-8"
                )
            except Exception:
                pass

        # 4.6 Парсим данные всех ЖК (ипотека, рассрочки, акции, текст)
        all_jk_names_set = set()
        for cells in rows:
            jk = cells[jk_idx].strip() if jk_idx is not None and jk_idx < len(cells) else ""
            if jk:
                all_jk_names_set.add(jk)

        existing_jk_data = {}
        if JK_DATA_FILE.exists():
            try:
                existing_jk_data = json.loads(JK_DATA_FILE.read_text(encoding="utf-8"))
            except:
                pass

        # Метки времени последнего парсинга ЖК
        JK_META_FILE = SCRIPT_DIR / "jk_parse_meta.json"
        jk_parse_meta = {}
        if JK_META_FILE.exists():
            try:
                jk_parse_meta = json.loads(JK_META_FILE.read_text(encoding="utf-8"))
            except:
                pass

        now_ts = time.time()
        JK_TTL = 7 * 24 * 3600  # 7 дней

        force = "--force" in sys.argv
        # Парсим только новые ЖК или те у которых данные старше 7 дней (или --force)
        jk_to_parse = sorted([
            jk for jk in all_jk_names_set
            if force
            or jk not in existing_jk_data
            or now_ts - jk_parse_meta.get(jk, 0) > JK_TTL
        ])
        jk_fresh = len(all_jk_names_set) - len(jk_to_parse)
        if jk_fresh:
            print(f"[•] ЖК: {jk_fresh} актуальных (< 7 дней), {len(jk_to_parse)} требуют обновления")
        # Получаем список ЖК с сайта один раз — используем и в шаге 4.6 и в шаге 6
        all_links = []
        if jk_to_parse or added_jk:
            all_links = get_object_links(driver)
            print(f"[•] Найдено ЖК на сайте: {len(all_links)}")

        if jk_to_parse:
            print(f"\n[•] Парсим данные ЖК ({len(jk_to_parse)} шт.)…")
            for jk_name in jk_to_parse:
                name, url = find_jk_link(jk_name, all_links)
                if url:
                    print(f"  [•] {jk_name}…", end=" ", flush=True)
                    driver.get(url)
                    wait_for_page_ready(driver)
                    for pct in [0.3, 0.6, 0.9, 1.0]:
                        driver.execute_script(f"window.scrollTo(0, document.body.scrollHeight * {pct});")
                        time.sleep(0.5)
                    driver.execute_script("window.scrollTo(0, 0);")
                    time.sleep(1)

                    is_new_jk = jk_name not in existing_jk_data or "--force" in sys.argv
                    _jkd_hint = {}
                    if JK_DATA_FILE.exists():
                        try:
                            _jkd_hint = json.loads(JK_DATA_FILE.read_text(encoding="utf-8"))
                        except Exception:
                            pass
                    jk_data = parse_jk_data(driver, jk_name, url, block_id_hint=_jkd_hint.get(jk_name, {}).get("block_id", ""))

                    # Для уже известных ЖК — сохраняем старые текст/карточки/фото
                    if not is_new_jk:
                        old = existing_jk_data[jk_name]
                        jk_data["about"]    = old.get("about",    jk_data["about"])
                        jk_data["features"] = old.get("features", jk_data["features"])

                    save_jk_data(jk_name, jk_data)
                    jk_parse_meta[jk_name] = now_ts
                    try:
                        JK_META_FILE.write_text(
                            json.dumps(jk_parse_meta, ensure_ascii=False, indent=2), encoding="utf-8"
                        )
                    except Exception:
                        pass
                    print(f"✓ (акций: {len(jk_data.get('promo', []))}, рассрочек: {len(jk_data.get('installments', []))})")
                else:
                    print(f"  [!] {jk_name}: не найден на сайте")


        print("\n[•] Сохраняем Excel…")
        xlsx_path = build_excel(headers, rows, date_str, layout_ids, apt_extra_map)
        print(f"[✓] Файл: {xlsx_path.name}")

        # Удаляем старые Excel файлы
        for old_file in SCRIPT_DIR.glob("trendagent_*.xlsx"):
            if old_file != xlsx_path:
                try:
                    old_file.unlink()
                    print(f"[•] Удалён старый файл: {old_file.name}")
                except PermissionError:
                    print(f"[!] Не удалось удалить {old_file.name} — закройте файл в Excel")

        # 6. Рендеры новых ЖК
        renders_downloaded = []
        if added_jk:
            print(f"\n[•] Новые ЖК — скачиваем рендеры ({len(added_jk)} шт.)…")
            print(f"  [•] Найдено ЖК на сайте: {len(all_links)}")
            for jk_name in sorted(added_jk):
                name, url = find_jk_link(jk_name, all_links)
                if url:
                    n = download_renders(driver, jk_name, url)
                    print(f"  [✓] {jk_name}: {n} фото")
                    renders_downloaded.append((jk_name, n))
                    u = upload_renders_to_s3(jk_name)
                    print(f"  [☁] {jk_name}: {u} фото → S3")
                    save_jk_data(jk_name, jk_data)
                    print(f"  [📋] {jk_name}: данные ЖК → json")
                else:
                    # Показываем похожие варианты из списка
                    jk_tr = translit(jk_name.lower())
                    hints = []
                    for name, lurl, slug in all_links:
                        score = sum(1 for w in jk_tr.split() if w in slug)
                        if score > 0:
                            hints.append(f"{name} ({slug})")
                    hint_str = " | ".join(hints[:2]) if hints else "—"
                    print(f"  [!] {jk_name}: ссылка не найдена. Похожие: {hint_str}")

        # 7. Отчёт
        print("\n" + "=" * 50)
        print("ОТЧЁТ")
        print("=" * 50)
        if prev_excel:
            print(f"  + Новых квартир:     {len(added_apts)}")
            print(f"  - Ушло квартир:      {len(removed_apts)}")
            print(f"  ≠ Изменились цены:   {len(changed_prices)}")
            if changed_prices:
                for key, (old_p, new_p) in list(changed_prices.items())[:10]:
                    # key = "ЖК__apt_id" или просто "apt_id"
                    display = key.replace("__", " / ") if "__" in key else key
                    print(f"      {display}: {old_p} → {new_p}")
                if len(changed_prices) > 10:
                    print(f"      ... и ещё {len(changed_prices)-10}")
            print(f"  + Новых ЖК:          {len(added_jk)}")
            if renders_downloaded:
                print(f"  📷 Рендеры:")
                for jk_name, n in renders_downloaded:
                    print(f"      {jk_name}: {n} фото")
        else:
            print(f"  Первый запуск: {len(rows)} квартир, {len(new_jk)} ЖК")
        print(f"  🏠 Планировок скачано: {newly_downloaded}")
        if newly_downloaded > 0:
            new_folder = LAYOUTS_DIR / "_new"
            new_files = sorted(new_folder.glob("*")) if new_folder.exists() else []
            for f in new_files:
                print(f"      📄 {f.name}")
            print(f"      📁 Папка: {new_folder}")
        elapsed = _time.time() - _start
        mins, secs = divmod(int(elapsed), 60)
        print(f"  ⏱  Время работы:      {mins} мин {secs} сек")
        print(f"  📄 Excel: {xlsx_path.name}")
        print("=" * 50)
        print("\n✅ Готово!")

        # ── Журнал прогона ────────────────────────────────────
        try:
            write_run("parser", {
                "cmd": "main",
                "duration_sec":  int(elapsed),
                "apts_total":    len(rows),
                "apts_added":    len(added_apts) if prev_excel else 0,
                "apts_removed":  len(removed_apts) if prev_excel else 0,
                "price_changes": len(changed_prices) if prev_excel else 0,
                "new_jk":        len(added_jk) if prev_excel else 0,
                "layouts_dl":    newly_downloaded,
                "excel":         xlsx_path.name,
                "first_run":     not prev_excel,
            })
            update_state()
            print("📝 PROJECT_STATE.md обновлён")
        except Exception as _log_e:
            print(f"[!] журнал состояния: {_log_e}")

    finally:
        driver.quit()


if __name__ == "__main__":
    import sys

    # ── Сброс планировки конкретной квартиры ──────────────────────
    if "--reset-layout" in sys.argv:
        idx = sys.argv.index("--reset-layout")
        if idx + 2 < len(sys.argv):
            _rl_jk  = sys.argv[idx + 1]
            _rl_apt = sys.argv[idx + 2]
            _rl_folder = LAYOUTS_DIR / safe_dirname(_rl_jk)
            _rl_fname  = f"{safe_filename(_rl_jk)}_{safe_filename(_rl_apt)}"
            _deleted = []
            for _ext in ("webp", "png", "jpg", "jpeg"):
                _f = _rl_folder / f"{_rl_fname}.{_ext}"
                if _f.exists():
                    _f.unlink()
                    _deleted.append(_f.name)
            if _deleted:
                print(f"[✓] Удалены файлы планировки: {', '.join(_deleted)}")
                print(f"[•] Теперь запустите парсер без флагов — он скачает планировку заново")
            else:
                print(f"[!] Файлы не найдены: {_rl_folder / _rl_fname}.*")
                print(f"[!] Проверьте название ЖК и номер квартиры")
            sys.exit(0)
        else:
            print("[!] Укажите ЖК и номер: --reset-layout \"Вилла Марина\" 15")
            sys.exit(1)

    if "--parse-jk-one" in sys.argv:
        # Парсим данные одного конкретного ЖК
        idx = sys.argv.index("--parse-jk-one")
        if idx + 1 >= len(sys.argv):
            print('[!] Укажите название ЖК: --parse-jk-one "SHEPILEVSKIY"')
            sys.exit(1)
        target_jk = sys.argv[idx + 1]
        print(f"[•] Парсим один ЖК: {target_jk}")
        driver = get_driver()
        try:
            ensure_auth(driver)
            all_links = get_object_links(driver)
            name, url = find_jk_link(target_jk, all_links)
            if not url:
                print(f"[!] ЖК «{target_jk}» не найден на сайте")
                driver.quit()
                sys.exit(1)
            print(f"[•] Найден: {name} → {url}")
            driver.get(url)
            wait_for_page_ready(driver)
            for pct in [0.3, 0.6, 0.9, 1.0]:
                driver.execute_script(f"window.scrollTo(0, document.body.scrollHeight * {pct});")
                time.sleep(0.5)
            driver.execute_script("window.scrollTo(0, 0);")
            time.sleep(1)
            existing_jk_data = {}
            if JK_DATA_FILE.exists():
                try:
                    existing_jk_data = json.loads(JK_DATA_FILE.read_text(encoding="utf-8"))
                except:
                    pass
            _jkd_hint = existing_jk_data.get(target_jk, {}).get("block_id", "")
            jk_data = parse_jk_data(driver, target_jk, url, block_id_hint=_jkd_hint)
            save_jk_data(target_jk, jk_data)
            print(f"[✓] {target_jk}: акций={len(jk_data.get('promo',[]))}, ипотека={len(jk_data.get('mortgage',[]))}, рассрочек={len(jk_data.get('installments',[]))}")
            for p in jk_data.get('promo', []):
                print(f"    • {p.get('name','')} | {p.get('value','')}")
        finally:
            driver.quit()
        sys.exit(0)

    if "--parse-jk-all" in sys.argv:
        # Парсим данные всех ЖК из Excel таблицы
        driver = get_driver()
        try:
            ensure_auth(driver)
            excel_path = get_prev_excel()
            if not excel_path:
                print("[!] Excel файл не найден")
                driver.quit()
                sys.exit(1)
            import pandas as pd
            df = pd.read_excel(excel_path)
            col_jk = next((c for c in df.columns if 'жк' in c.lower() or c == 'ЖК'), None)
            if not col_jk:
                print("[!] Колонка ЖК не найдена в Excel")
            else:
                all_jk = sorted(set(str(v).strip() for v in df[col_jk].dropna() if str(v).strip() not in ('', 'nan')))
                print(f"[•] Найдено ЖК в таблице: {len(all_jk)}")
                all_links = get_object_links(driver)
                print(f"[•] Найдено ЖК на сайте: {len(all_links)}")

                existing_jk_data = {}
                if JK_DATA_FILE.exists():
                    try:
                        existing_jk_data = json.loads(JK_DATA_FILE.read_text(encoding="utf-8"))
                    except:
                        pass

                for n, jk_name in enumerate(all_jk, 1):
                    name, url = find_jk_link(jk_name, all_links)
                    if not url:
                        print(f"  [{n}/{len(all_jk)}] {jk_name}: не найден на сайте")
                        continue
                    print(f"  [{n}/{len(all_jk)}] {jk_name}…", end=" ", flush=True)
                    driver.get(url)
                    wait_for_page_ready(driver)
                    for pct in [0.3, 0.6, 0.9, 1.0]:
                        driver.execute_script(f"window.scrollTo(0, document.body.scrollHeight * {pct});")
                        time.sleep(0.5)
                    driver.execute_script("window.scrollTo(0, 0);")
                    time.sleep(1)
                    is_new = jk_name not in existing_jk_data or "--force" in sys.argv
                    _jkd_hint = {}
                    if JK_DATA_FILE.exists():
                        try:
                            _jkd_hint = json.loads(JK_DATA_FILE.read_text(encoding="utf-8"))
                        except Exception:
                            pass
                    jk_data = parse_jk_data(driver, jk_name, url, block_id_hint=_jkd_hint.get(jk_name, {}).get("block_id", ""))
                    if not is_new:
                        old = existing_jk_data[jk_name]
                        jk_data["about"]    = old.get("about",    jk_data["about"])
                        jk_data["features"] = old.get("features", jk_data["features"])
                    save_jk_data(jk_name, jk_data)
                    print(f"✓ (акций: {len(jk_data.get('promo',[]))}, рассрочек: {len(jk_data.get('installments',[]))}, ипотека: {len(jk_data.get('mortgage',[]))})")
        finally:
            driver.quit()
    elif "--parse-jk" in sys.argv:
        # Только парсим данные ЖК (текст, карточки, акции) без рендеров
        jk_list = [a for a in sys.argv[1:] if not a.startswith("--")]
        driver = get_driver()
        try:
            ensure_auth(driver)
            all_links = get_object_links(driver)
            print(f"[•] Найдено ЖК на сайте: {len(all_links)}")
            if not jk_list:
                print("[!] Укажите название ЖК: --parse-jk \"Название ЖК\"")
            else:
                # Если первый аргумент — URL, второй может быть именем ЖК
                jk_url  = next((a for a in jk_list if a.startswith("http")), None)
                jk_name_override = next((a for a in jk_list if not a.startswith("http")), None)
                loop_list = [(jk_url, jk_name_override)] if jk_url else [(None, a) for a in jk_list]

                for jk_url_arg, jk_name_arg in loop_list:
                    if jk_url_arg:
                        url = jk_url_arg
                        jk_name = jk_name_arg if jk_name_arg else jk_url_arg.split("/object/")[-1].split("/")[0].replace("-", " ").upper()
                    else:
                        jk_name = jk_name_arg
                        name, url = find_jk_link(jk_name, all_links)
                    if url:
                        print(f"[•] URL: {url}")
                        print(f"[•] Парсим данные: {jk_name}...")
                        driver.get(url)
                        wait_for_page_ready(driver)
                        # Скроллим до конца чтобы подгрузить все блоки включая акции
                        for pct in [0.3, 0.6, 0.9, 1.0]:
                            driver.execute_script(f"window.scrollTo(0, document.body.scrollHeight * {pct});")
                            time.sleep(0.5)
                        driver.execute_script("window.scrollTo(0, 0);")
                        time.sleep(1)
                        _jkd_hint = {}
                        if JK_DATA_FILE.exists():
                            try:
                                _jkd_hint = json.loads(JK_DATA_FILE.read_text(encoding="utf-8"))
                            except Exception:
                                pass
                        jk_data = parse_jk_data(driver, jk_name, url, block_id_hint=_jkd_hint.get(jk_name, {}).get("block_id", ""))
                        print(f"  Текст: {len(jk_data['about'])} символов")
                        print(f"  Карточки: {len(jk_data['features'])} шт.")
                        print(f"  Акции: {len(jk_data['promo'])} шт.")
                        for p in jk_data['promo']:
                            desc = p.get('description', '').strip()
                            print(f"    • {p['name']} — {p['value']} ({p['conditions']})")
                            if desc:
                                print(f"      ↳ {desc[:120]}")
                        save_jk_data(jk_name, jk_data)
                        print(f"[✓] Сохранено в jk_data.json")
                    else:
                        print(f"[!] {jk_name}: не найден")
        finally:
            driver.quit()
    elif "--fix-layout" in sys.argv:
        # Перекачать планировку одной квартиры: --fix-layout "ЖК" "apt_id"
        args = [a for a in sys.argv[1:] if not a.startswith("--")]
        if len(args) < 2:
            print('[!] Укажите: --fix-layout "ЖК" "apt_id"')
        else:
            jk_name, apt_id = args[0], args[1]
            driver = get_driver()
            try:
                ensure_auth(driver)
                expand_all(driver)
                row_els = driver.find_elements(By.CSS_SELECTOR, "tr.apartments-table__row")
                soup = BeautifulSoup(driver.page_source, "html.parser")
                headers, rows = parse_table(soup)
                id_idx = headers.index("id / №") if "id / №" in headers else None
                jk_idx  = headers.index("ЖК") if "ЖК" in headers else None
                found = False
                for i, cells in enumerate(rows):
                    cid = cells[id_idx].strip() if id_idx is not None else ""
                    cjk = cells[jk_idx].strip() if jk_idx is not None else ""
                    if cid == apt_id and cjk == jk_name:
                        print(f"[•] Скачиваем планировку: {jk_name} / {apt_id}…", end=" ", flush=True)
                        # Удаляем старый файл
                        folder = LAYOUTS_DIR / safe_dirname(jk_name)
                        fname = f"{safe_filename(jk_name)}_{safe_filename(apt_id)}"
                        for old_f in folder.glob(f"{fname}.*"):
                            old_f.unlink()
                        ok, _ = download_layout_via_click(driver, row_els[i], apt_id, jk_name)
                        print("✓" if ok else "не удалось")
                        found = True
                        break
                if not found:
                    print(f"[!] Квартира {apt_id} в ЖК {jk_name} не найдена")
            finally:
                driver.quit()
    elif "--check-renders" in sys.argv:
        # Проверяем каких ЖК нет в папке renders
        import pandas as pd
        excel_path = get_prev_excel()
        if not excel_path:
            print("[!] Excel не найден")
        else:
            df = pd.read_excel(excel_path)
            col_jk = next((c for c in df.columns if c == 'ЖК'), None)
            all_jk = sorted(set(str(v).strip() for v in df[col_jk].dropna() if str(v).strip() not in ('', 'nan')))
            print(f"[•] ЖК в таблице: {len(all_jk)}")
            missing = []
            for jk in all_jk:
                folder = RENDERS_DIR / safe_dirname(jk)
                files = list(folder.glob("*")) if folder.exists() else []
                if not files:
                    missing.append(jk)
                    print(f"  [!] Нет рендеров: {jk}")
            print(f"\n[•] Итого без рендеров: {len(missing)}")
            if missing:
                print("\nЗапусти:")
                for jk in missing:
                    print(f'  --fix-renders "{jk}"')
    elif "--fix-renders-all" in sys.argv:
        # Перекачать ВСЕ ЖК с сайта и залить в S3
        RENDERS_DIR.mkdir(exist_ok=True)
        driver = get_driver()
        try:
            ensure_auth(driver)
            all_links = get_object_links(driver)
            print(f"[•] Найдено ЖК на сайте: {len(all_links)}")
            # Фильтруем — только записи с нормальным коротким именем
            clean_links = [(n.split(chr(10))[0].strip(), u, s) for n, u, s in all_links if n.split(chr(10))[0].strip()]
            print(f"[•] Скачиваем рендеры для всех {len(clean_links)} ЖК...")
            for jk_name, url, slug in sorted(clean_links):
                # Удаляем старую папку чтобы перекачать в полном качестве
                folder = RENDERS_DIR / safe_dirname(jk_name)
                if folder.exists():
                    shutil.rmtree(folder)
                n = download_renders(driver, jk_name, url)
                print(f"  [✓] {jk_name}: {n} фото")
                if n > 0:
                    u = upload_renders_to_s3(jk_name)
                    print(f"  [☁] {jk_name}: {u} фото → S3")
                else:
                    print(f"  [!] {jk_name}: фото не найдены")
        finally:
            driver.quit()
    elif "--fix-renders" in sys.argv:
        jk_list = [a for a in sys.argv[1:] if not a.startswith("--")]
        RENDERS_DIR.mkdir(exist_ok=True)
        driver = get_driver()
        try:
            ensure_auth(driver)
            all_links = get_object_links(driver)
            print(f"[•] Найдено ЖК на сайте: {len(all_links)}")
            if not jk_list:
                print("[•] Все ЖК на сайте:")
                for n, u, s in sorted(all_links):
                    print(f"  - {n} ({s})")
            else:
                print(f"[•] Скачиваем рендеры для {len(jk_list)} ЖК...")
                for jk_name in jk_list:
                    name, url = find_jk_link(jk_name, all_links)
                    if url:
                        n = download_renders(driver, jk_name, url)
                        print(f"  [✓] {jk_name}: {n} фото")
                        u = upload_renders_to_s3(jk_name)
                        print(f"  [☁] {jk_name}: {u} фото → S3")
                    else:
                        print(f"  [!] {jk_name}: не найден")
        finally:
            driver.quit()
    elif "--check-data" in sys.argv:
        # Показывает какие ЖК в jk_data.json не имеют данных (ипотека/рассрочки/акции)
        if not JK_DATA_FILE.exists():
            print("[!] jk_data.json не найден")
        else:
            all_data = json.loads(JK_DATA_FILE.read_text(encoding="utf-8"))
            print(f"[•] Всего ЖК в jk_data.json: {len(all_data)}")
            missing = []
            for jk, data in sorted(all_data.items()):
                mort  = len(data.get("mortgage",      []))
                inst  = len(data.get("installments",  []))
                promo = len(data.get("promo",         []))
                flags = []
                if mort  == 0: flags.append("ипотека")
                if inst  == 0: flags.append("рассрочки")
                if promo == 0: flags.append("акции")
                if flags:
                    missing.append(jk)
                    print(f"  [!] {jk:<40} нет: {', '.join(flags)}")
                else:
                    print(f"  [✓] {jk:<40} mort={mort} inst={inst} promo={promo}")
            print(f"\n[•] ЖК без данных: {len(missing)} из {len(all_data)}")
            if missing:
                print("\nЗапусти для дозаполнения:")
                print(f"  python {sys.argv[0]} --fix-missing-data")
    elif "--fix-missing-data" in sys.argv:
        # Перепарсит только ЖК у которых нет ипотеки, рассрочек или акций
        force = "--force" in sys.argv
        if not JK_DATA_FILE.exists():
            print("[!] jk_data.json не найден — запусти полный парсинг")
            sys.exit(1)
        all_data = json.loads(JK_DATA_FILE.read_text(encoding="utf-8"))
        to_fix = []
        for jk, data in all_data.items():
            mort  = len(data.get("mortgage",     []))
            inst  = len(data.get("installments", []))
            promo = len(data.get("promo",        []))
            if mort == 0 or inst == 0 or promo == 0:
                to_fix.append(jk)
        if not to_fix:
            print("[✓] Все ЖК уже имеют данные — ничего не нужно")
            sys.exit(0)
        print(f"[•] ЖК без данных: {len(to_fix)}")
        for jk in to_fix:
            d = all_data[jk]
            flags = []
            if not d.get("mortgage"):     flags.append("ипотека")
            if not d.get("installments"): flags.append("рассрочки")
            if not d.get("promo"):        flags.append("акции")
            print(f"  - {jk}  (нет: {', '.join(flags)})")
        print()
        driver = get_driver()
        try:
            ensure_auth(driver)
            all_links = get_object_links(driver)
            print(f"[•] Найдено ЖК на сайте: {len(all_links)}")
            for n, jk_name in enumerate(to_fix, 1):
                name, url = find_jk_link(jk_name, all_links)
                if not url:
                    print(f"  [{n}/{len(to_fix)}] {jk_name}: не найден на сайте — пропускаем")
                    continue
                print(f"  [{n}/{len(to_fix)}] {jk_name}…", end=" ", flush=True)
                driver.get(url)
                wait_for_page_ready(driver)
                for pct in [0.3, 0.6, 0.9, 1.0]:
                    driver.execute_script(f"window.scrollTo(0, document.body.scrollHeight * {pct});")
                    time.sleep(0.5)
                driver.execute_script("window.scrollTo(0, 0);")
                time.sleep(1)
                _jkd_hint = {}
                if JK_DATA_FILE.exists():
                    try:
                        _jkd_hint = json.loads(JK_DATA_FILE.read_text(encoding="utf-8"))
                    except Exception:
                        pass
                jk_data = parse_jk_data(driver, jk_name, url, block_id_hint=_jkd_hint.get(jk_name, {}).get("block_id", ""))
                # Сохраняем старые about/features если они уже есть
                old = all_data.get(jk_name, {})
                if not force:
                    if old.get("about"):    jk_data["about"]    = old["about"]
                    if old.get("features"): jk_data["features"] = old["features"]
                save_jk_data(jk_name, jk_data)
                mort  = len(jk_data.get("mortgage",     []))
                inst  = len(jk_data.get("installments", []))
                promo = len(jk_data.get("promo",        []))
                print(f"✓  (ипотека: {mort}, рассрочки: {inst}, акции: {promo})")
        finally:
            driver.quit()
    elif "--debug-apts" in sys.argv:
        # Дебаг базовых цен по квартирам конкретного ЖК
        # Использование: --debug-apts "Кронфорт"
        jk_filter = next((a for a in sys.argv[1:] if not a.startswith("--")), None)
        if not jk_filter:
            print("[!] Укажите название ЖК: --debug-apts \"Кронфорт\"")
            sys.exit(1)
        driver = get_driver()
        try:
            ensure_auth(driver)
            expand_all(driver)
            soup = BeautifulSoup(driver.page_source, "html.parser")
            headers, rows = parse_table(soup)
            id_idx = headers.index("id / №") if "id / №" in headers else None
            jk_idx = headers.index("ЖК")     if "ЖК"     in headers else None
            price_idx = next((i for i, h in enumerate(headers) if "100%" in str(h)), None)
            filtered = [(i, cells) for i, cells in enumerate(rows)
                        if jk_idx is not None and jk_filter.lower() in cells[jk_idx].lower()]
            print(f"[•] Найдено квартир {jk_filter}: {len(filtered)}")
            for n, (i, cells) in enumerate(filtered, 1):
                apt_id  = cells[id_idx].strip()  if id_idx  is not None else str(i)
                price   = cells[price_idx].strip() if price_idx is not None else ""
                print(f"\n  [{n}/{len(filtered)}] {apt_id}  (цена 100%: {price})")
                row_els = driver.find_elements(By.CSS_SELECTOR, "tr.apartments-table__row")
                if i >= len(row_els):
                    print("    строка не найдена")
                    continue
                tabs_before = set(driver.window_handles)
                driver.execute_script("arguments[0].click();", row_els[i])
                for _ in range(10):
                    time.sleep(0.5)
                    if set(driver.window_handles) - tabs_before:
                        break
                new_tabs = set(driver.window_handles) - tabs_before
                if not new_tabs:
                    print("    вкладка не открылась")
                    continue
                driver.switch_to.window(new_tabs.pop())
                wait_for_page_ready(driver)
                time.sleep(1.5)
                rows_data = driver.execute_script(
                    "var rows=document.querySelectorAll('.apartment-passport__row');"
                    "var result={};"
                    "rows.forEach(function(row){"
                    "var text=row.innerText.trim();"
                    "var parts=text.split('\\n');"
                    "if(parts.length>=2)result[parts[0].trim()]=parts.slice(1).join(' ').trim();"
                    "});"
                    "return result;"
                ) or {}
                bp   = rows_data.get("Базовая цена", "—")
                p100 = rows_data.get("Цена за м² при 100% оплате", "—")
                print(f"    Базовая цена:       {bp!r}")
                print(f"    Цена 100% (карточка): {p100!r}")
                print(f"    Все поля: {list(rows_data.keys())}")
                driver.close()
                driver.switch_to.window(driver.window_handles[0])
                time.sleep(0.5)
        finally:
            driver.quit()
    elif "--fix-base-prices" in sys.argv:
        # Перепарсивает базовые цены для всех квартир и пересохраняет Excel
        driver = get_driver()
        try:
            ensure_auth(driver)
            expand_all(driver)
            soup = BeautifulSoup(driver.page_source, "html.parser")
            headers, rows = parse_table(soup)
            if not rows:
                print("[!] Таблица пуста")
                driver.quit()
                sys.exit(1)
            id_idx = headers.index("id / №") if "id / №" in headers else None
            jk_idx = headers.index("ЖК")     if "ЖК"     in headers else None
            print(f"[•] Квартир: {len(rows)}, собираем базовые цены…")
            apt_extra_map = {}
            FIELDS = {
                "Цена за м² при 100% оплате": "цена_за_м2",
                "Базовая цена":               "базовая_цена",
                "Цена за м² при базовой цене":"цена_за_м2_базовая",
                "Окна":                       "окна",
                "Вид из окна":                "вид_из_окна",
                "Видовая квартира":            "видовая_квартира",
                "Срок сдачи":                 "срок_сдачи",
                "Дата старта продаж":         "дата_старта_продаж",
                "Класс недвижимости":         "класс_недвижимости",
                "Высота потолков":            "высота_потолков",
                "Старт продаж, цена":         "старт_продаж_цена",
            }
            only_missing = "--only-missing" in sys.argv
            # Если --only-missing, читаем текущий Excel и находим квартиры без доп.данных
            missing_keys = set()
            if only_missing:
                import re as _re
                prev_xl = get_prev_excel()
                if prev_xl:
                    import pandas as _pd
                    _df = _pd.read_excel(prev_xl)
                    _extra_cols = ['Базовая цена', 'Окна', 'Вид из окна', 'Высота потолков']
                    _col_jk = next((c for c in _df.columns if c == 'ЖК'), None)
                    _col_id = next((c for c in _df.columns if 'id' in str(c).lower() or '№' in str(c)), None)
                    for _, row in _df.iterrows():
                        jk_v  = str(row.get(_col_jk, '')).strip()
                        id_v  = str(row.get(_col_id, '')).strip()
                        if not jk_v or not id_v: continue
                        all_empty = all(_pd.isna(row.get(c)) or str(row.get(c, '')).strip() in ('', 'nan')
                                        for c in _extra_cols if c in _df.columns)
                        if all_empty:
                            missing_keys.add(f"{jk_v}_{id_v}")
                    print(f"[•] Квартир без доп.данных: {len(missing_keys)}")
            for row_idx, cells in enumerate(rows):
                apt_id = cells[id_idx].strip() if id_idx is not None and id_idx < len(cells) else ""
                jk     = cells[jk_idx].strip() if jk_idx is not None and jk_idx < len(cells) else ""
                if not apt_id or not jk:
                    continue
                if only_missing and f"{jk}_{apt_id}" not in missing_keys:
                    continue
                print(f"  [{row_idx+1}/{len(rows)}] {jk} / {apt_id}…", end=" ", flush=True)
                try:
                    row_els = driver.find_elements(By.CSS_SELECTOR, "tr.apartments-table__row")
                    if row_idx >= len(row_els):
                        print("не найдена")
                        continue
                    tabs_before = set(driver.window_handles)
                    driver.execute_script("arguments[0].click();", row_els[row_idx])
                    for _ in range(10):
                        time.sleep(0.5)
                        if set(driver.window_handles) - tabs_before:
                            break
                    new_tabs = set(driver.window_handles) - tabs_before
                    if not new_tabs:
                        print("вкладка не открылась")
                        continue
                    driver.switch_to.window(new_tabs.pop())
                    wait_for_page_ready(driver)
                    time.sleep(2)
                    # Ждём появления паспорта квартиры
                    for _ in range(10):
                        chk = driver.execute_script(
                            "var r=document.querySelectorAll('.apartment-passport__row');"
                            "return r.length > 0 ? r[0].innerText.trim() : '';"
                        ) or ''
                        if chk and '\u200c' not in chk and len(chk) > 3:
                            break
                        time.sleep(0.5)
                    rows_data = driver.execute_script(
                        "var rows=document.querySelectorAll('.apartment-passport__row');"
                        "var result={};"
                        "rows.forEach(function(row){"
                        "var text=row.innerText.trim();"
                        "var parts=text.split('\\n');"
                        "if(parts.length>=2)result[parts[0].trim()]=parts.slice(1).join(' ').trim();"
                        "});"
                        "return result;"
                    ) or {}
                    # Фильтруем мусорные ключи (невидимые символы)
                    rows_data = {k: v for k, v in rows_data.items() 
                                 if k.strip() and '\u200c' not in k and len(k.strip()) > 1}
                    apt_extra = {v: rows_data[k] for k, v in FIELDS.items() if k in rows_data}
                    apt_extra_map[f"{jk}_{apt_id}"] = apt_extra
                    bp = apt_extra.get("базовая_цена", "—")
                    print(f"✓  базовая: {bp}")
                    if bp == "—" and rows_data:
                        print(f"    [DEBUG] Поля паспорта: {list(rows_data.keys())}")
                        for k2, v2 in rows_data.items():
                            print(f"      {k2!r}: {v2!r}")
                    driver.close()
                    driver.switch_to.window(driver.window_handles[0])
                    time.sleep(0.5)
                except Exception as e:
                    print(f"[!] {e}")
                    try:
                        driver.close()
                        driver.switch_to.window(driver.window_handles[0])
                    except:
                        pass
            # Пересохраняем Excel
            from datetime import datetime
            date_str = datetime.now().strftime("%Y-%m-%d_%H-%M")
            # Берём layout_ids и СУЩЕСТВУЮЩИЕ доп.данные из предыдущего Excel
            prev = get_prev_excel()
            layout_ids = set()
            if prev:
                import pandas as _pd2, re as _re2
                wb_old = load_workbook(prev)
                ws_old = wb_old.active
                hdrs_old = [ws_old.cell(1, c).value for c in range(1, ws_old.max_column + 1)]
                plan_idx = next((i for i, h in enumerate(hdrs_old) if h == "Планировка"), None)
                jk_idx_old  = next((i for i, h in enumerate(hdrs_old) if h == "ЖК"), None)
                id_idx_old  = next((i for i, h in enumerate(hdrs_old) if h == "id / №"), None)
                EXTRA_MAP = {
                    "Цена за м²":           "цена_за_м2",
                    "Базовая цена":         "базовая_цена",
                    "Цена за м² (баз.)":    "цена_за_м2_базовая",
                    "Окна":                 "окна",
                    "Вид из окна":          "вид_из_окна",
                    "Срок сдачи (карточка)":"срок_сдачи",
                    "Дата старта продаж":   "дата_старта_продаж",
                    "Класс":                "класс_недвижимости",
                    "Высота потолков":      "высота_потолков",
                    "Старт продаж, цена":   "старт_продаж_цена",
                }
                extra_col_idx = {col: i for i, col in enumerate(hdrs_old) if col in EXTRA_MAP}
                for row in ws_old.iter_rows(min_row=2, values_only=True):
                    if plan_idx is not None and row[plan_idx]:
                        layout_ids.add(str(row[plan_idx]).split("_")[-1].split(".")[0])
                    if jk_idx_old is None or id_idx_old is None: continue
                    jk_v = str(row[jk_idx_old] or '').strip()
                    id_v = str(row[id_idx_old] or '').strip()
                    if not jk_v or not id_v: continue
                    key = f"{jk_v}_{id_v}"
                    # Берём существующие доп.данные только если в новом парсе их нет
                    if key not in apt_extra_map:
                        old_extra = {}
                        for col, field in EXTRA_MAP.items():
                            ci = extra_col_idx.get(col)
                            if ci is not None and row[ci] and str(row[ci]).strip() not in ('', 'nan'):
                                old_extra[field] = str(row[ci]).strip()
                        if old_extra:
                            apt_extra_map[key] = old_extra
            xlsx_path = build_excel(headers, rows, date_str, layout_ids, apt_extra_map)
            print(f"\n[✓] Excel сохранён: {xlsx_path.name}")
            # Удаляем старые
            for old_file in SCRIPT_DIR.glob("trendagent_*.xlsx"):
                if old_file != xlsx_path:
                    try: old_file.unlink(); print(f"[•] Удалён: {old_file.name}")
                    except: pass
        finally:
            driver.quit()
    else:
        main()

# --- Автосинк в git-репо ---
try:
    from git_sync import git_sync
    git_sync("parser run")
except Exception as e:
    print(f"git_sync skipped: {e}")
