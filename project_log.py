#!/usr/bin/env python3
"""
project_log.py — единый журнал состояния проекта.

Поддерживает три файла рядом со скриптами:
  • runs.jsonl        — append-only история запусков (1 строка = 1 прогон)
  • PROJECT_STATE.md  — живой снапшот состояния (перезаписывается)
  • DECISIONS.md      — ручные архитектурные решения (не трогается)

Использование из парсера/генератора:
    from project_log import write_run, update_state

    # В конце main():
    write_run("parser", {
        "cmd": "main",
        "duration_sec": int(time.time() - _start),
        "apts_total": len(rows),
        "apts_added": len(added_apts),
        ...
    })
    update_state()  # пересоздаёт PROJECT_STATE.md

Сохраняет ручные секции в PROJECT_STATE.md между маркерами:
    <!-- KEEP_START:todo -->
    ... ваш текст ...
    <!-- KEEP_END:todo -->
"""

from __future__ import annotations

import json
import re
from datetime import datetime
from pathlib import Path

# Папка с проектом — там же где этот модуль
BASE_DIR = Path(__file__).parent

RUNS_FILE  = BASE_DIR / "runs.jsonl"
STATE_FILE = BASE_DIR / "PROJECT_STATE.md"

# Сколько последних запусков показывать в STATE
RUNS_IN_STATE = 8


# ──────────────────────────────────────────────────────────────
# WRITE_RUN — append одной строки JSON в runs.jsonl
# ──────────────────────────────────────────────────────────────
def write_run(kind: str, data: dict) -> None:
    """
    Дописывает запись о прогоне в runs.jsonl.
    kind: "parser" | "generator" | другое (например "fix-renders")
    data: любые сериализуемые поля. Время добавится автоматически.

    Никогда не падает — все ошибки гасятся, чтобы не ронять основной скрипт.
    """
    try:
        entry = {
            "ts": datetime.now().isoformat(timespec="seconds"),
            "kind": kind,
            **data,
        }
        with open(RUNS_FILE, "a", encoding="utf-8") as f:
            f.write(json.dumps(entry, ensure_ascii=False) + "\n")
    except Exception as e:
        print(f"[project_log] не удалось записать прогон: {e}")


def read_runs(limit: int | None = None) -> list[dict]:
    """Читает runs.jsonl и возвращает список записей (новые в конце)."""
    if not RUNS_FILE.exists():
        return []
    try:
        with open(RUNS_FILE, encoding="utf-8") as f:
            lines = [ln.strip() for ln in f if ln.strip()]
        records = []
        for ln in lines:
            try:
                records.append(json.loads(ln))
            except Exception:
                continue
        if limit:
            records = records[-limit:]
        return records
    except Exception as e:
        print(f"[project_log] не удалось прочитать runs.jsonl: {e}")
        return []


# ──────────────────────────────────────────────────────────────
# СТАТИСТИКА — собирает текущее состояние из файлов проекта
# ──────────────────────────────────────────────────────────────
def _latest_excel() -> Path | None:
    files = sorted(BASE_DIR.glob("trendagent_*.xlsx"), reverse=True)
    return files[0] if files else None


def _count_layouts() -> int:
    """Считает webp файлы в layouts/{ЖК}/"""
    layouts = BASE_DIR / "layouts"
    if not layouts.exists():
        return 0
    n = 0
    for sub in layouts.iterdir():
        if sub.is_dir() and sub.name != "_new":
            n += sum(1 for f in sub.iterdir() if f.suffix.lower() == ".webp")
    return n


def _count_renders() -> tuple[int, int]:
    """Возвращает (число ЖК с рендерами, общее число файлов)."""
    renders = BASE_DIR / "renders"
    if not renders.exists():
        return 0, 0
    jk_with = 0
    total = 0
    for sub in renders.iterdir():
        if not sub.is_dir():
            continue
        files = [f for f in sub.iterdir()
                 if f.suffix.lower() in (".jpg", ".jpeg", ".png", ".webp")]
        if files:
            jk_with += 1
            total += len(files)
    return jk_with, total


def _read_jk_data() -> dict:
    f = BASE_DIR / "jk_data.json"
    if not f.exists():
        return {}
    try:
        return json.loads(f.read_text(encoding="utf-8"))
    except Exception:
        return {}


def _read_excel_summary() -> dict:
    """Считает квартиры, ЖК, покрытие планировками из последнего xlsx."""
    xlsx = _latest_excel()
    if not xlsx:
        return {"excel": None, "apts": 0, "jks": 0, "with_layout": 0}
    try:
        from openpyxl import load_workbook
        wb = load_workbook(xlsx, read_only=True, data_only=True)
        ws = wb.active
        headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]

        def col_idx(*variants):
            for i, h in enumerate(headers):
                if not h:
                    continue
                for v in variants:
                    if v.lower() in str(h).lower():
                        return i
            return None

        jk_i  = col_idx("ЖК")
        lay_i = col_idx("планировк")
        apts = 0
        jks = set()
        with_layout = 0
        for row in ws.iter_rows(min_row=2, values_only=True):
            apts += 1
            if jk_i is not None and row[jk_i]:
                jks.add(str(row[jk_i]).strip())
            if lay_i is not None and row[lay_i] and str(row[lay_i]).strip():
                with_layout += 1
        wb.close()
        return {
            "excel": xlsx.name,
            "apts": apts,
            "jks": len(jks),
            "with_layout": with_layout,
        }
    except Exception as e:
        return {"excel": xlsx.name, "apts": 0, "jks": 0, "with_layout": 0,
                "error": str(e)}


def _jk_data_stats() -> dict:
    data = _read_jk_data()
    if not data:
        return {"total": 0, "with_promo": 0, "with_inst": 0, "with_mort": 0}
    return {
        "total": len(data),
        "with_promo": sum(1 for v in data.values() if v.get("promo")),
        "with_inst":  sum(1 for v in data.values() if v.get("installments")),
        "with_mort":  sum(1 for v in data.values() if v.get("mortgage")),
    }


# ──────────────────────────────────────────────────────────────
# UPDATE_STATE — перегенерирует PROJECT_STATE.md
# ──────────────────────────────────────────────────────────────

# Маркеры ручных секций — содержимое внутри них переносится из старого файла
KEEP_PATTERN = re.compile(
    r"<!--\s*KEEP_START:(\w+)\s*-->(.*?)<!--\s*KEEP_END:\1\s*-->",
    re.DOTALL,
)


def _read_keep_blocks() -> dict[str, str]:
    """Достаёт ручные секции из текущего PROJECT_STATE.md (если есть)."""
    if not STATE_FILE.exists():
        return {}
    try:
        content = STATE_FILE.read_text(encoding="utf-8")
        return {m.group(1): m.group(2) for m in KEEP_PATTERN.finditer(content)}
    except Exception:
        return {}


def _format_run(r: dict) -> str:
    """Одна строка истории прогонов."""
    ts = r.get("ts", "")[:16].replace("T", " ")
    kind = r.get("kind", "?")
    dur = r.get("duration_sec")
    dur_s = f"{dur//60}м{dur%60:02d}с" if isinstance(dur, int) and dur >= 60 \
            else (f"{dur}с" if dur is not None else "")
    extras = []
    if "apts_added"  in r: extras.append(f"+{r['apts_added']}")
    if "apts_removed" in r: extras.append(f"-{r['apts_removed']}")
    if "price_changes" in r: extras.append(f"≠{r['price_changes']}")
    if "new_jk" in r and r["new_jk"]: extras.append(f"новых ЖК: {r['new_jk']}")
    if "layouts_dl" in r and r["layouts_dl"]: extras.append(f"планировок: {r['layouts_dl']}")
    if "share_pages_uploaded" in r and r["share_pages_uploaded"]:
        extras.append(f"share: {r['share_pages_uploaded']}")
    if "errors" in r and r["errors"]:
        extras.append(f"⚠️ ошибок: {len(r['errors'])}")
    extra_s = "  " + ", ".join(extras) if extras else ""
    return f"- `{ts}` **{kind}** ({dur_s}){extra_s}"


def update_state() -> None:
    """
    Перегенерирует PROJECT_STATE.md.
    Сохраняет ручные секции (внутри маркеров KEEP_START/KEEP_END).
    """
    try:
        keep = _read_keep_blocks()
        excel_info = _read_excel_summary()
        jk_stats   = _jk_data_stats()
        layouts_n  = _count_layouts()
        jk_with_renders, renders_total = _count_renders()
        runs = read_runs(limit=RUNS_IN_STATE)

        last_parser = next((r for r in reversed(runs) if r.get("kind") == "parser"), None)
        last_gen    = next((r for r in reversed(runs) if r.get("kind") == "generator"), None)

        def _last_line(r: dict | None) -> str:
            if not r:
                return "_не запускался_"
            ts = r.get("ts", "")[:16].replace("T", " ")
            dur = r.get("duration_sec", 0)
            dur_s = f"{dur//60} мин {dur%60:02d} сек" if dur >= 60 else f"{dur} сек"
            return f"`{ts}` ({dur_s})"

        # ── Дефолты для ручных секций (если ещё не заполнены) ────
        keep_todo = keep.get("todo") or "\n- [ ] _добавьте сюда ваши задачи_\n"
        keep_recent = keep.get("recent") or "\n- _здесь можно описывать недавние изменения_\n"
        keep_known_issues = keep.get("known_issues") or "\n- _здесь — известные баги и обходы_\n"

        # ── История прогонов ─────────────────────────────────────
        history_lines = [_format_run(r) for r in reversed(runs)] or ["_прогонов ещё не было_"]
        history = "\n".join(history_lines)

        # ── Сборка markdown ──────────────────────────────────────
        excel_name = excel_info.get("excel") or "_не найден_"
        coverage = (f"{excel_info['with_layout']} / {excel_info['apts']}"
                    if excel_info.get("apts") else "—")

        md = f"""# TrendAgent Terrace — состояние проекта

_Обновлено: {datetime.now().strftime("%Y-%m-%d %H:%M")}_

## 📊 Снапшот данных

- **Excel:** `{excel_name}`
- **Квартир в базе:** {excel_info.get('apts', 0)}
- **ЖК:** {excel_info.get('jks', 0)} (в `jk_data.json`: {jk_stats['total']})
- **Планировок WebP:** {layouts_n} (в Excel с привязкой: {coverage})
- **Рендеры:** {jk_with_renders} ЖК с фото, всего {renders_total} файлов
- **Покрытие данными ЖК:** акции — {jk_stats['with_promo']}, рассрочки — {jk_stats['with_inst']}, ипотека — {jk_stats['with_mort']}

## ⚙️ Последние прогоны

- **Парсер:** {_last_line(last_parser)}
- **Генератор:** {_last_line(last_gen)}

## 🏗 Архитектура (на текущий момент)

- **Парсер** (`trendagent_parser_*.py`) → Selenium → spb.trendagent.ru
  → `trendagent_<дата>.xlsx`, `jk_data.json`, кэши
  → Яндекс S3: `royaltyplace/terrace/layouts/`, `renders/`, `renders-thumb/`
- **Генератор** (`Generate_filter_*.py`) → читает xlsx + jk_data.json + Google Sheets
  → `index.html`, `filter.css`, `filter.js`, `filter-data.js`, `share/*.html`
  → Яндекс S3: `royaltyplace/terrace/`
- **Сайт:** [terrace-royaltyplace.ru](https://terrace-royaltyplace.ru) — статика из S3

## 📌 TODO

<!-- KEEP_START:todo -->{keep_todo}<!-- KEEP_END:todo -->

## 🐛 Известные баги и обходы

<!-- KEEP_START:known_issues -->{keep_known_issues}<!-- KEEP_END:known_issues -->

## 📝 Последние изменения в коде

<!-- KEEP_START:recent -->{keep_recent}<!-- KEEP_END:recent -->

## 🕒 История прогонов (последние {RUNS_IN_STATE})

{history}

---
_Эти секции автогенерятся: «Снапшот данных», «Последние прогоны», «История прогонов».
Секции внутри `KEEP_START/KEEP_END` редактируются вручную и сохраняются между обновлениями.
Полная история прогонов — в `runs.jsonl`._
"""
        STATE_FILE.write_text(md, encoding="utf-8")
    except Exception as e:
        print(f"[project_log] не удалось обновить PROJECT_STATE.md: {e}")


# ──────────────────────────────────────────────────────────────
# CLI — для ручного запуска: python project_log.py
# ──────────────────────────────────────────────────────────────
if __name__ == "__main__":
    update_state()
    print(f"✅ {STATE_FILE.name} обновлён")
    runs = read_runs(limit=5)
    print(f"📚 Последние {len(runs)} прогонов в {RUNS_FILE.name}:")
    for r in runs:
        print("   " + _format_run(r))
